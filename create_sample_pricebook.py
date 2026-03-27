"""
create_sample_pricebook.py

Generates a minimal sample "Pricelist WorkBook.xlsx" that matches
the column layout expected by bom/pricelist_parser.py.

Run:  python create_sample_pricebook.py
Output: Sample_Pricelist_WorkBook.xlsx  (in the same folder)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl import Workbook

H = Font(bold=True)
GREY = PatternFill("solid", fgColor="D9D9D9")

wb = Workbook()
wb.remove(wb.active)   # remove default sheet


# ─────────────────────────────────────────────────────────────────
# 1. RATES  (three column groups side by side)
#
# Parser reads (0-indexed columns):
#   Group A Raw Material  : col 1=item  col 2=price  col 3=prev
#   Group B Bought Out    : col 9=item  col 10=price col 12=prev
#   Group C ENCON Purchase: col 15=item col 19=price col 20=prev
#
# openpyxl columns are 1-indexed, so add 1 everywhere.
# ─────────────────────────────────────────────────────────────────

ws = wb.create_sheet("Rates")

def hdr(ws, row, col, val):
    c = ws.cell(row=row, column=col, value=val)
    c.font = H
    c.fill = GREY

# Group A header (col 2,3,4 → 0-idx 1,2,3)
hdr(ws, 1, 2, "RAW MATERIAL ITEM")
hdr(ws, 1, 3, "PRICE")
hdr(ws, 1, 4, "PREVIOUS PRICE")

# Group B header (col 10,11,13 → 0-idx 9,10,12)
hdr(ws, 1, 10, "BOUGHT OUT ITEMS")
hdr(ws, 1, 11, "PRICE")
hdr(ws, 1, 13, "PREVIOUS PRICE")

# Group C header (col 16,20,21 → 0-idx 15,19,20)
hdr(ws, 1, 16, "ENCON PURCHASE PRICE")
hdr(ws, 1, 20, "PRICE")
hdr(ws, 1, 21, "PREVIOUS PRICE")

raw_materials = [
    ("MS Plate (6mm)",          85,   80),
    ("MS Plate (8mm)",          85,   80),
    ("SS Pipe 50 NB",           320,  300),
    ("MS Pipe 25 NB",           95,   90),
    ("Ceramic Fiber Roll",      4200, 4000),
    ("Refractory Bricks",       45,   42),
    ("MS Angle 50x50x5",        80,   75),
    ("MS Channel 100x50",       92,   88),
    ("Paint (primer)",          250,  230),
    ("Nuts & Bolts (kg)",       120,  110),
]

bought_out = [
    ("Ball Valve 20 NB",        850,   800),
    ("Ball Valve 25 NB",        1100,  1000),
    ("Butterfly Valve 80 NB",   4500,  4200),
    ("Butterfly Valve 100 NB",  5500,  5200),
    ("Solenoid Valve 15 NB",    2200,  2000),
    ("Pressure Gauge with TNV", 1800,  1700),
    ("Pressure Gauge with NV",  1600,  1500),
    ("Pressure Switch Low",     3200,  3000),
    ("Pressure Switch High + Low", 3500, 3300),
    ("Compensator 80 NB",       2800,  2600),
    ("Thermocouple",            1500,  1400),
    ("Compensating Lead",       180,   170),
    ("Limit Switches",          2500,  2300),
    ("Temperature Transmitter", 4500,  4200),
    ("P.PID",                   8500,  8000),
    ("Ratio Controller",        12000, 11000),
    ("Cable for Ignition Transformer", 350, 320),
    ("Hydraulic Power Pack & Cylinder", 85000, 80000),
    ("Flexible Hose 20 NB",     1200,  1100),
    ("Flexible Hose 15 NB",     900,   850),
]

encon_purchase = [
    ("Ignition Transformer",    3500,  3200),
    ("Burner Control Unit",     12000, 11000),
    ("UV Sensor with Air Jacket", 4500, 4200),
    ("ENCON-PB (NG/LPG) - 100 KW", 18000, 17000),
    ("Pressure Regulating Valve", 2800, 2600),
]

for i, (item, price, prev) in enumerate(raw_materials, start=2):
    ws.cell(row=i, column=2, value=item)
    ws.cell(row=i, column=3, value=price)
    ws.cell(row=i, column=4, value=prev)

for i, (item, price, prev) in enumerate(bought_out, start=2):
    ws.cell(row=i, column=10, value=item)
    ws.cell(row=i, column=11, value=price)
    ws.cell(row=i, column=13, value=prev)

for i, (item, price, prev) in enumerate(encon_purchase, start=2):
    ws.cell(row=i, column=16, value=item)
    ws.cell(row=i, column=20, value=price)
    ws.cell(row=i, column=21, value=prev)


# ─────────────────────────────────────────────────────────────────
# 2. HPU
#
# Row 1 (0-idx 0): "3 KW Costing"  at col A, "6 KW Costing" at col F
# Row 2 (0-idx 1): variant names   "Duplex 1"  "Simplex 1"
# Row 3 (0-idx 2): column headers  "Items" "Qty" "Unit" "Rate" "Amount"
# Rows 4+        : data
# ─────────────────────────────────────────────────────────────────

ws = wb.create_sheet("HPU")

# 3 KW block starts at col A (col 1), Duplex 1 at A, Simplex 1 at F
ws.cell(row=1, column=1, value="3 KW Costing").font = H
ws.cell(row=1, column=6, value="6 KW Costing").font = H

ws.cell(row=2, column=1, value="Duplex 1")
ws.cell(row=2, column=6, value="Duplex 1")

for base_col in (1, 6):
    ws.cell(row=3, column=base_col,   value="Items")
    ws.cell(row=3, column=base_col+1, value="Qty")
    ws.cell(row=3, column=base_col+2, value="Unit")
    ws.cell(row=3, column=base_col+3, value="Rate")
    ws.cell(row=3, column=base_col+4, value="Amount")

hpu_3kw = [
    ("Motor 3 KW",          1, "Nos", 8500,  8500),
    ("Pump (Gear)",         1, "Nos", 12000, 12000),
    ("Tank (50L)",          1, "Nos", 6500,  6500),
    ("Pressure Gauge",      2, "Nos", 800,   1600),
    ("Relief Valve",        1, "Nos", 2200,  2200),
    ("Solenoid Valve",      2, "Nos", 1800,  3600),
    ("Hoses & Fittings",    1, "Set", 3200,  3200),
]

hpu_6kw = [
    ("Motor 6 KW",          1, "Nos", 14500, 14500),
    ("Pump (Gear)",         1, "Nos", 16000, 16000),
    ("Tank (100L)",         1, "Nos", 9500,  9500),
    ("Pressure Gauge",      2, "Nos", 800,   1600),
    ("Relief Valve",        1, "Nos", 2500,  2500),
    ("Solenoid Valve",      2, "Nos", 1800,  3600),
    ("Hoses & Fittings",    1, "Set", 4500,  4500),
]

for r, (item, qty, unit, rate, amt) in enumerate(hpu_3kw, start=4):
    ws.cell(row=r, column=1, value=item)
    ws.cell(row=r, column=2, value=qty)
    ws.cell(row=r, column=3, value=unit)
    ws.cell(row=r, column=4, value=rate)
    ws.cell(row=r, column=5, value=amt)

for r, (item, qty, unit, rate, amt) in enumerate(hpu_6kw, start=4):
    ws.cell(row=r, column=6,  value=item)
    ws.cell(row=r, column=7,  value=qty)
    ws.cell(row=r, column=8,  value=unit)
    ws.cell(row=r, column=9,  value=rate)
    ws.cell(row=r, column=10, value=amt)


# ─────────────────────────────────────────────────────────────────
# 3. BURNER  (burner_pricelist_master)
#
# Section header row (single cell)
# "BURNER SIZE" header row
# Data rows: burner_size, Component1_price, Component2_price ...
# ─────────────────────────────────────────────────────────────────

ws = wb.create_sheet("BURNER")

row = 1
ws.cell(row=row, column=1, value="NATURAL GAS BURNER").font = H;  row += 1
ws.cell(row=row, column=1, value="BURNER SIZE")
ws.cell(row=row, column=2, value="BURNER ALONE")
ws.cell(row=row, column=3, value="WITH PILOTS")
ws.cell(row=row, column=4, value="COMPLETE SET")
for col in range(1, 5):
    ws.cell(row=row, column=col).font = H
row += 1

ng_burners = [
    ("4A",  45000, 52000, 68000),
    ("6A",  58000, 66000, 85000),
    ("8A",  72000, 82000, 105000),
    ("10A", 88000, 100000, 128000),
]
for size, b, p, c in ng_burners:
    ws.cell(row=row, column=1, value=f"ENCON {size}")
    ws.cell(row=row, column=2, value=b)
    ws.cell(row=row, column=3, value=p)
    ws.cell(row=row, column=4, value=c)
    row += 1

row += 1
ws.cell(row=row, column=1, value="LPG GAS BURNER").font = H;  row += 1
ws.cell(row=row, column=1, value="BURNER SIZE")
ws.cell(row=row, column=2, value="BURNER ALONE")
ws.cell(row=row, column=3, value="WITH PILOTS")
ws.cell(row=row, column=4, value="COMPLETE SET")
for col in range(1, 5):
    ws.cell(row=row, column=col).font = H
row += 1

lpg_burners = [
    ("4A",  47000, 54000, 70000),
    ("6A",  60000, 68000, 88000),
]
for size, b, p, c in lpg_burners:
    ws.cell(row=row, column=1, value=f"ENCON {size}")
    ws.cell(row=row, column=2, value=b)
    ws.cell(row=row, column=3, value=p)
    ws.cell(row=row, column=4, value=c)
    row += 1


# ─────────────────────────────────────────────────────────────────
# 4. Blower  (blower_pricelist_master)
# ─────────────────────────────────────────────────────────────────

ws = wb.create_sheet("Blower")

row = 1
ws.cell(row=row, column=1, value="MEDIUM PRESSURE").font = H;  row += 1
ws.cell(row=row, column=1, value="MODEL")
ws.cell(row=row, column=2, value="HP")
ws.cell(row=row, column=3, value="PRESSURE (mmWC)")
ws.cell(row=row, column=4, value="FLOW (Nm3/hr)")
ws.cell(row=row, column=5, value="BASIC PRICE")
ws.cell(row=row, column=6, value="WITH MOTOR")
for col in range(1, 7):
    ws.cell(row=row, column=col).font = H
row += 1

mp_blowers = [
    ("ENCON EBB-3",  3,  800, 250,  28000, 38000),
    ("ENCON EBB-5",  5,  900, 400,  35000, 48000),
    ("ENCON EBB-7",  7,  1000, 600, 44000, 60000),
    ("ENCON EBB-10", 10, 1100, 900, 55000, 75000),
]
for model, hp, pres, flow, basic, with_motor in mp_blowers:
    ws.cell(row=row, column=1, value=model)
    ws.cell(row=row, column=2, value=hp)
    ws.cell(row=row, column=3, value=pres)
    ws.cell(row=row, column=4, value=flow)
    ws.cell(row=row, column=5, value=basic)
    ws.cell(row=row, column=6, value=with_motor)
    row += 1

row += 1
ws.cell(row=row, column=1, value="HIGH PRESSURE").font = H;  row += 1
ws.cell(row=row, column=1, value="MODEL")
ws.cell(row=row, column=2, value="HP")
ws.cell(row=row, column=3, value="PRESSURE (mmWC)")
ws.cell(row=row, column=4, value="FLOW (Nm3/hr)")
ws.cell(row=row, column=5, value="BASIC PRICE")
ws.cell(row=row, column=6, value="WITH MOTOR")
for col in range(1, 7):
    ws.cell(row=row, column=col).font = H
row += 1

hp_blowers = [
    ("ENCON EBH-5",  5,  2000, 300, 42000, 58000),
    ("ENCON EBH-7",  7,  2200, 480, 52000, 72000),
    ("ENCON EBH-10", 10, 2500, 700, 65000, 88000),
]
for model, hp, pres, flow, basic, with_motor in hp_blowers:
    ws.cell(row=row, column=1, value=model)
    ws.cell(row=row, column=2, value=hp)
    ws.cell(row=row, column=3, value=pres)
    ws.cell(row=row, column=4, value=flow)
    ws.cell(row=row, column=5, value=basic)
    ws.cell(row=row, column=6, value=with_motor)
    row += 1


# ─────────────────────────────────────────────────────────────────
# 5. Horizontal  (horizontal_master)
#
# Model header row → HORIZONTAL LADLE PREHEATER 10 T
# Then particular rows: [s.no], particular, qty_text, amount
# ─────────────────────────────────────────────────────────────────

ws = wb.create_sheet("Horizontal")

row = 1
for tons, ms_kg, ms_rate, cf_rolls, panel, trolley, pipeline in [
    (10,  1096, 151.2,  9,  126000, 46200,  124200),
    (20,  1300, 151.2,  10, 126000, 52500,  158372),
    (30,  1354, 151.2,  11, 147000, 58000,  168000),
]:
    ws.cell(row=row, column=1,
            value=f"HORIZONTAL LADLE PREHEATER {tons} T").font = H
    row += 1
    particulars = [
        ("MS STRUCTURE",           f"{ms_kg} KGS",     round(ms_kg * ms_rate)),
        ("CERAMIC FIBER",          f"{cf_rolls} ROLLS", cf_rolls * 4200),
        ("CONTROL PANEL",          "1 SET",             panel),
        ("TROLLEY DRIVE MECHANISM","",                  trolley),
        ("PIPELINE & FITTINGS",    "",                  pipeline),
    ]
    for sno, (part, qty_txt, amt) in enumerate(particulars, start=1):
        ws.cell(row=row, column=1, value=sno)
        ws.cell(row=row, column=2, value=part)
        ws.cell(row=row, column=3, value=qty_txt)
        ws.cell(row=row, column=4, value=amt)
        row += 1
    row += 1  # blank separator


# ─────────────────────────────────────────────────────────────────
# 6. Vertical  (vertical_master)
#
# Each model occupies a block of columns (5-wide).
# Row with "VERTICAL" = model header.
# Then particular rows.
# ─────────────────────────────────────────────────────────────────

ws = wb.create_sheet("Vertical")

models = [
    (10,  1096, 151.2,  9,  105000, 124929),
    (20,  1300, 151.2,  10, 115500, 158372),
    (30,  1354, 151.2,  11, 147000, 168000),
]

for mi, (tons, ms_kg, ms_rate, cf_rolls, panel, pipeline_swirl) in enumerate(models):
    base_col = mi * 5 + 1
    r = 1
    ws.cell(row=r, column=base_col,
            value=f"VERTICAL LADLE PREHEATER {tons} T").font = H
    r += 1
    particulars = [
        ("MS STRUCTURE",                f"{ms_kg} KGS",     round(ms_kg * ms_rate)),
        ("CERAMIC FIBER",               f"{cf_rolls} ROLLS", cf_rolls * 4200),
        ("CONTROL PANEL",               "1 SET",             panel),
        ("SWIRLING MECH, PIPELINE & FITTINGS", "",           pipeline_swirl),
    ]
    for sno, (part, qty_txt, amt) in enumerate(particulars, start=1):
        ws.cell(row=r, column=base_col,   value=sno)
        ws.cell(row=r, column=base_col+1, value=part)
        ws.cell(row=r, column=base_col+2, value=qty_txt)
        ws.cell(row=r, column=base_col+3, value=amt)
        r += 1


# ─────────────────────────────────────────────────────────────────
# 7. Gas Burner parts  (gas_burner_parts_master)
# ─────────────────────────────────────────────────────────────────

ws = wb.create_sheet("Gas Burner")

row = 1
ws.cell(row=row, column=1, value="NATURAL GAS BURNER PARTS").font = H; row += 1
for sno, (part, qty, unit, rate, amt) in enumerate([
    ("Nozzle",        1, "NOS", 2500, 2500),
    ("Diffuser",      1, "NOS", 1800, 1800),
    ("Mixing Tube",   1, "NOS", 3200, 3200),
    ("Air Register",  1, "NOS", 4500, 4500),
], start=1):
    ws.cell(row=row, column=1, value=sno)
    ws.cell(row=row, column=2, value=part)
    ws.cell(row=row, column=3, value=qty)
    ws.cell(row=row, column=4, value=unit)
    ws.cell(row=row, column=5, value=rate)
    ws.cell(row=row, column=6, value=amt)
    row += 1


# ─────────────────────────────────────────────────────────────────
# 8. Oil Burner parts  (oil_burner_parts_master)
# ─────────────────────────────────────────────────────────────────

ws = wb.create_sheet(" Oil Burner")   # note leading space (matches init_db.py)

row = 1
ws.cell(row=row, column=1, value="OIL BURNER PARTS").font = H; row += 1
for sno, (part, qty, unit, rate, amt) in enumerate([
    ("Nozzle (Oil)",   1, "NOS", 1800, 1800),
    ("Filter",         1, "NOS", 2200, 2200),
    ("Pump (Oil)",     1, "NOS", 8500, 8500),
], start=1):
    ws.cell(row=row, column=1, value=sno)
    ws.cell(row=row, column=2, value=part)
    ws.cell(row=row, column=3, value=qty)
    ws.cell(row=row, column=4, value=unit)
    ws.cell(row=row, column=5, value=rate)
    ws.cell(row=row, column=6, value=amt)
    row += 1


# ─────────────────────────────────────────────────────────────────
# 9. HV Oil Burner parts  (hv_oil_burner_parts_master)
# ─────────────────────────────────────────────────────────────────

ws = wb.create_sheet("HV  Oil Burner")   # double space matches init_db.py

row = 1
ws.cell(row=row, column=1, value="HV OIL BURNER PARTS").font = H; row += 1
for sno, (part, qty, unit, rate, amt) in enumerate([
    ("HV Nozzle",      1, "NOS", 2800, 2800),
    ("HV Filter",      1, "NOS", 3200, 3200),
], start=1):
    ws.cell(row=row, column=1, value=sno)
    ws.cell(row=row, column=2, value=part)
    ws.cell(row=row, column=3, value=qty)
    ws.cell(row=row, column=4, value=unit)
    ws.cell(row=row, column=5, value=rate)
    ws.cell(row=row, column=6, value=amt)
    row += 1


# ─────────────────────────────────────────────────────────────────
# 10. Recuperator  (recuperator_master)
# ─────────────────────────────────────────────────────────────────

ws = wb.create_sheet("Recuperator")

row = 1
ws.cell(row=row, column=1, value="RECUPERATOR TYPE A").font = H; row += 1
ws.cell(row=row, column=1, value="MODEL")
ws.cell(row=row, column=2, value="CAPACITY (KW)")
ws.cell(row=row, column=3, value='TUBE DIA (in)')
ws.cell(row=row, column=4, value="PRICE")
for col in range(1, 5):
    ws.cell(row=row, column=col).font = H
row += 1
for model, cap, dia, price in [
    ("100", 100, 1.5, 45000),
    ("200", 200, 2.0, 68000),
    ("300", 300, 2.5, 92000),
]:
    ws.cell(row=row, column=1, value=model)
    ws.cell(row=row, column=2, value=cap)
    ws.cell(row=row, column=3, value=dia)
    ws.cell(row=row, column=4, value=price)
    row += 1


# ─────────────────────────────────────────────────────────────────
# 11. GAIL GAS Burner  (gail_gas_burner_master)
# ─────────────────────────────────────────────────────────────────

ws = wb.create_sheet("GAIL GAS Burner")

row = 1
ws.cell(row=row, column=1, value="GAIL GAS BURNER PRICE LIST").font = H; row += 1
ws.cell(row=row, column=1, value="BURNER SIZE")
ws.cell(row=row, column=2, value="BURNER ALONE")
ws.cell(row=row, column=3, value="WITH PILOTS")
for col in range(1, 4):
    ws.cell(row=row, column=col).font = H
row += 1
for size, b, p in [
    ("ENCON G-4A",  48000, 56000),
    ("ENCON G-6A",  62000, 72000),
    ("ENCON G-8A",  78000, 90000),
]:
    ws.cell(row=row, column=1, value=size)
    ws.cell(row=row, column=2, value=b)
    ws.cell(row=row, column=3, value=p)
    row += 1


# ─────────────────────────────────────────────────────────────────
# 12. Rad Heat (TATA)  (rad_heat_tata_master)
# ─────────────────────────────────────────────────────────────────

ws = wb.create_sheet("Rad Heat (TATA)")

row = 1
ws.cell(row=row, column=1, value="PRICE LIST OF RAD-HEAT (TATA)").font = H; row += 1
ws.cell(row=row, column=1, value="MODEL")
ws.cell(row=row, column=2, value="OUTPUT KW")
ws.cell(row=row, column=3, value="GAS LPG (M3/hr)")
ws.cell(row=row, column=4, value="GAS NG (M3/hr)")
ws.cell(row=row, column=5, value="PRICE WITH SS TUBING")
for col in range(1, 6):
    ws.cell(row=row, column=col).font = H
row += 1
for model, kw, lpg, ng, price in [
    ("ARE-100", 100, 12, 10, 85000),
    ("ARE-200", 200, 24, 20, 145000),
    ("ARE-300", 300, 36, 30, 195000),
]:
    ws.cell(row=row, column=1, value=model)
    ws.cell(row=row, column=2, value=kw)
    ws.cell(row=row, column=3, value=lpg)
    ws.cell(row=row, column=4, value=ng)
    ws.cell(row=row, column=5, value=price)
    row += 1

row += 1
ws.cell(row=row, column=1, value="SPARES PRICE LIST").font = H; row += 1
for spare, price in [
    ("Tube Assembly", 12000),
    ("Reflector",      4500),
    ("Igniter",        2800),
]:
    ws.cell(row=row, column=1, value=spare)
    ws.cell(row=row, column=2, value=price)
    row += 1


# ─────────────────────────────────────────────────────────────────
# 13. Rad Heat  (rad_heat_master)
# ─────────────────────────────────────────────────────────────────

ws = wb.create_sheet("Rad Heat")

row = 1
ws.cell(row=row, column=1, value="PRICE LIST OF RAD-HEAT").font = H; row += 1
ws.cell(row=row, column=1, value="MODEL")
ws.cell(row=row, column=2, value="OUTPUT KW")
ws.cell(row=row, column=3, value="GAS LPG (M3/hr)")
ws.cell(row=row, column=4, value="GAS NG (M3/hr)")
ws.cell(row=row, column=5, value="PRICE WITH SS TUBING")
for col in range(1, 6):
    ws.cell(row=row, column=col).font = H
row += 1
for model, kw, lpg, ng, price in [
    ("ARE-100", 100, 12, 10, 82000),
    ("ARE-200", 200, 24, 20, 140000),
    ("ARE-300", 300, 36, 30, 188000),
]:
    ws.cell(row=row, column=1, value=model)
    ws.cell(row=row, column=2, value=kw)
    ws.cell(row=row, column=3, value=lpg)
    ws.cell(row=row, column=4, value=ng)
    ws.cell(row=row, column=5, value=price)
    row += 1

row += 1
ws.cell(row=row, column=1, value="SPARES PRICE LIST").font = H; row += 1
for spare, price in [
    ("Tube Assembly", 11000),
    ("Reflector",      4200),
]:
    ws.cell(row=row, column=1, value=spare)
    ws.cell(row=row, column=2, value=price)
    row += 1


# ─────────────────────────────────────────────────────────────────
# SAVE
# ─────────────────────────────────────────────────────────────────

out = "Sample_Pricelist_WorkBook.xlsx"
wb.save(out)
print(f"OK Created: {out}")
print()
print("Sheets included:")
for s in wb.sheetnames:
    print(f"  - {s}")
print()
print("Upload this file at the Price Master page > 'Upload Pricelist' button.")
print("All 13 DB tables will be refreshed in one shot.")
