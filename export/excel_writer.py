"""
FINAL frozen Excel writer
Matches legacy Excel logic EXACTLY
"""

import pandas as pd
import io
from typing import Dict
from openpyxl import load_workbook
from export.calculation_sheet import write_calculation_sheet


# -------------------------------------------------
# LEGACY EXCLUSION RULES
# -------------------------------------------------
BOUGHT_OUT_EXCLUDE_ITEMS = {
    "COMPENSATOR",
    "PRESSURE GAUGE WITH TNV",
    "RATIO CONTROLLER",
}


def write_excel(
    buffer,
    sheets: Dict[str, pd.DataFrame],
    burner_inputs,
    burner_results,
    pipe_results,
):
    """
    Writes Excel with:
    - BOM sheet
    - Cost Summary
    - Calculation sheet (via openpyxl)
    """

    # -------------------------------------------------
    # 1️⃣ Write BOM & Cost Summary (xlsxwriter)
    # -------------------------------------------------
    with pd.ExcelWriter(buffer, engine="xlsxwriter") as writer:
        workbook = writer.book

        # Identify BOM sheet
        bom_sheet_name = next(
            name for name in sheets.keys() if name != "Cost Summary"
        )

        bom_df = sheets[bom_sheet_name].copy()

        # Ensure numeric totals
        bom_df["TOTAL"] = pd.to_numeric(
            bom_df["TOTAL"], errors="coerce"
        ).fillna(0)

        # ---------- ENCON TOTAL ----------
        encon_total = bom_df.loc[
            bom_df["MEDIA"] == "ENCON ITEMS",
            "TOTAL",
        ].sum()

        # ---------- BOUGHT OUT TOTAL ----------
        bought_out_total = bom_df.loc[
            (bom_df["MEDIA"] != "ENCON ITEMS")
            & (~bom_df["ITEM NAME"].isin(BOUGHT_OUT_EXCLUDE_ITEMS)),
            "TOTAL",
        ].sum()

        # ---------- Final BOM ----------
        final_bom_df = bom_df.copy()

        final_bom_df.to_excel(
            writer,
            sheet_name=bom_sheet_name,
            index=False,
        )

        worksheet = writer.sheets[bom_sheet_name]

        header_fmt = workbook.add_format({
            "bold": True,
            "border": 1,
            "align": "center",
            "valign": "vcenter",
        })

        for col_num, col_name in enumerate(final_bom_df.columns):
            worksheet.write(0, col_num, col_name, header_fmt)
            worksheet.set_column(col_num, col_num, 22)

        # ---------- Cost Summary ----------
        if "Cost Summary" in sheets:
            sheets["Cost Summary"].to_excel(
                writer,
                sheet_name="Cost Summary",
                index=False,
            )

    # -------------------------------------------------
    # 2️⃣ Inject Calculation Sheet (openpyxl)
    # -------------------------------------------------
    buffer.seek(0)
    wb = load_workbook(buffer)

    if burner_inputs is not None:
        ws = wb.create_sheet("Calculation", 0)

        write_calculation_sheet(
            ws,
            burner_inputs,
            burner_results,
            pipe_results,
        )

    # -------------------------------------------------
    # 🔥 CRITICAL FIX: Save into fresh buffer
    # -------------------------------------------------
    new_buffer = io.BytesIO()
    wb.save(new_buffer)
    new_buffer.seek(0)

    # Replace original buffer safely
    buffer.truncate(0)
    buffer.seek(0)
    buffer.write(new_buffer.getvalue())
    buffer.seek(0)