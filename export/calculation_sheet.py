# export/calculation_sheet.py
"""
Engineering-style Calculation Sheet layout
(Burner + NG Pipe + Air Pipe)
NO calculations here — presentation only
"""

from openpyxl.styles import Font, PatternFill, Border, Side


# -----------------------------
# STYLES
# -----------------------------
BOLD = Font(bold=True)

YELLOW = PatternFill(
    fill_type="solid",
    fgColor="FFFF00"
)

THIN = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


# -----------------------------
# MAIN WRITER
# -----------------------------
def write_calculation_sheet(
    ws,
    burner_inputs,
    burner_results,
    pipe_results,
):
    """
    Writes the Calculation sheet layout exactly like Excel

    Parameters
    ----------
    ws : openpyxl worksheet
    burner_inputs : BurnerInputs
    burner_results : BurnerResults
    pipe_results : PipeResults
    """

    # =================================================
    # HEADERS
    # =================================================
    ws["B2"] = "Burner Size Calculation"
    ws["B2"].font = BOLD

    ws["H2"] = "NG PIPE SIZE CALCULATION"
    ws["H2"].font = BOLD

    ws["H13"] = "AIR PIPE SIZE CALCULATION"
    ws["H13"].font = BOLD

    # =================================================
    # LEFT BLOCK – BURNER CALCULATION
    # =================================================
    rows = [
        ("Ti", burner_inputs.Ti, "°C"),
        ("Tf", burner_inputs.Tf, "°C"),
        ("Actual Refractory Weight", burner_inputs.refractory_weight, "Kg"),
        ("MG Fuel CV", burner_inputs.fuel_cv, "Kcal/Nm³"),

        ("Average Temperature Rise", burner_results.avg_temp_rise, "°C"),
        ("Time Taken", burner_inputs.time_taken_hr, "Hr"),

        ("Fuel Consumption", burner_results.fuel_consumption_nm3, "Nm³"),
        ("Calculated Firing Rate", burner_results.calculated_firing_rate_nm3hr, "Nm³/hr"),
        ("10% Extra Firing Rate", burner_results.extra_firing_rate_nm3hr, "Nm³/hr"),
        ("Final Firing Rate", burner_results.extra_firing_rate_nm3hr, "Nm³/hr"),

        ("Air Qty", burner_results.air_qty_nm3hr, "Nm³/hr"),
        ("CFM", burner_results.cfm, "CFM"),
        ("Blower Size", burner_results.blower_hp, "HP"),
    ]

    start_row = 4
    for i, (label, value, unit) in enumerate(rows):
        r = start_row + i

        ws[f"B{r}"] = label
        ws[f"C{r}"] = value
        ws[f"D{r}"] = unit

        ws[f"C{r}"].fill = YELLOW

        for col in ("B", "C", "D"):
            ws[f"{col}{r}"].border = THIN

    # =================================================
    # NG PIPE SIZE CALCULATION
    # =================================================
    ws["H4"] = "Q = FLOW (Nm³/hr)"
    ws["I4"] = round(burner_results.extra_firing_rate_nm3hr, 2)

    ws["H5"] = "V = VELOCITY (m/s)"
    ws["I5"] = 12.7
    ws["I5"].fill = YELLOW

    ws["H6"] = "d = Inner Dia (mm)"
    ws["I6"] = round(pipe_results.ng_pipe_inner_dia_mm, 2)
    ws["I6"].fill = YELLOW

    ws["I7"] = "125 NB"

    # =================================================
    # AIR PIPE SIZE CALCULATION
    # =================================================
    ws["H15"] = "Q = FLOW (Nm³/hr)"
    ws["I15"] = round(burner_results.air_qty_nm3hr, 2)

    ws["H16"] = "V = VELOCITY (m/s)"
    ws["I16"] = 15
    ws["I16"].fill = YELLOW

    ws["H17"] = "d = Inner Dia (mm)"
    ws["I17"] = round(pipe_results.air_pipe_inner_dia_mm, 2)
    ws["I17"].fill = YELLOW

    ws["I18"] = "300 NB"
