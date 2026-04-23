"""
bom/pricelist_parser.py

Parses all sheets of the ENCON Pricelist WorkBook into SQLite tables.
Used by both init_db.py and the /api/upload-pricelist endpoint.
"""
import re
import math
import sqlite3
import pandas as pd


# ─────────────────────────────────────────────────────────────────
# SS Pipe price calculation
# Formula matches Excel:  E = C * G
#   G (kg/mtr) = π/4 × (OD² - ID²) × 7850 / 1,000,000
#   E (Rs/mtr) = price_per_kg × kg_per_mtr
# ─────────────────────────────────────────────────────────────────

def _ss_pipe_price_per_mtr(item_name: str, price_per_kg: float):
    """
    Parse OD and wall from name like 'SS Pipe 304 60 X 3mm'
    and return price per metre.  Returns None if name doesn't match.
    """
    m = re.search(r'(\d+)\s*[Xx]\s*(\d+(?:\.\d+)?)\s*mm', item_name, re.I)
    if not m:
        return None
    od_mm   = float(m.group(1))
    wall_mm = float(m.group(2))
    id_mm   = od_mm - 2 * wall_mm
    kg_per_mtr = math.pi * (od_mm**2 - id_mm**2) * 7850 / (4 * 1_000_000)
    return round(price_per_kg * kg_per_mtr, 4)


def safe_float(v):
    try:
        return float(str(v).replace(",", "").strip())
    except Exception:
        return None


def _find_sheet(xl, keyword):
    """Case-insensitive sheet name search. Exact match wins over substring."""
    kl = keyword.lower()
    # 1. Exact match (strips surrounding whitespace from sheet name)
    for s in xl.sheet_names:
        if s.strip().lower() == kl:
            return s
    # 2. Substring match
    for s in xl.sheet_names:
        if kl in s.strip().lower():
            return s
    return None


# ─────────────────────────────────────────────────────────────────
# 1. RATES → component_price_master
# ─────────────────────────────────────────────────────────────────

def parse_rates(xl, conn):
    sheet = _find_sheet(xl, "rates")
    if sheet is None:
        return {"skipped": "Rates sheet not found"}

    df = xl.parse(sheet, header=None)

    SKIP_LOWER = {
        "price", "item", "previous", "bought out items",
        "encon purchase price", "specification", "s.no",
        "price list data", "price june", "out items",
    }

    def is_header(v):
        if not isinstance(v, str):
            return False
        return any(kw in v.lower() for kw in SKIP_LOWER)

    def clean_num(v):
        try:
            f = float(v)
            return f if f > 0 else None
        except (TypeError, ValueError):
            return None

    def clean_text(v):
        if not isinstance(v, str):
            return None
        s = v.strip()
        # Normalize: collapse multiple spaces, strip trailing spaces inside parens
        s = re.sub(r'\s+', ' ', s)
        s = re.sub(r'\(\s+', '(', s)
        s = re.sub(r'\s+\)', ')', s)
        # Remove quotes around single letters (M.S. Tube "B" -> M.S. Tube B)
        s = s.replace('"', '')
        return s if len(s) >= 2 and not s.replace(".", "").replace(",", "").isdigit() else None

    rows = []  # tuples of (item, category, unit, price, prev, excel_row, excel_col)
    for pandas_idx, row in df.iterrows():
        excel_row = int(pandas_idx) + 1  # 1-based Excel row

        # Group A: Raw Material — item col 1, price col 2 → Excel col C = 3
        item  = clean_text(row.iloc[1] if len(row) > 1 else None)
        price = clean_num(row.iloc[2]  if len(row) > 2 else None)
        prev  = clean_num(row.iloc[3]  if len(row) > 3 else None)
        if item and price and not is_header(item):
            # SS Pipe: compute price per metre from geometry formula
            # Skip raw "SS Pipe 304 60 X 3mm" rows — the "(per mtr)" computed row will be kept instead
            if re.search(r'ss\s*pipe', item, re.I) and re.search(r'\d+\s*[Xx]\s*\d+\s*mm', item, re.I) and 'per mtr' not in item.lower():
                continue
            unit = "kg" if price <= 500 else "nos"
            if "per mtr" in item.lower() or "(per mtr)" in item.lower():
                unit = "mtr"
            rows.append((item, "Raw Material", unit, price, prev or price, excel_row, 3))

        # Group B: Bought Out — item col 9, price col 10 → Excel col K = 11
        item  = clean_text(row.iloc[9]  if len(row) > 9  else None)
        price = clean_num(row.iloc[10]  if len(row) > 10 else None)
        prev  = clean_num(row.iloc[12]  if len(row) > 12 else None)
        if item and price and not is_header(item):
            rows.append((item, "Bought Out", "nos", price, prev or price, excel_row, 11))

        # Group C: ENCON Purchase — item col 15, price col 19 → Excel col T = 20
        item  = clean_text(row.iloc[15] if len(row) > 15 else None)
        price = clean_num(row.iloc[19]  if len(row) > 19 else None)
        prev  = clean_num(row.iloc[20]  if len(row) > 20 else None)
        if item and price and not is_header(item):
            rows.append((item, "ENCON Purchase", "nos", price, prev or price, excel_row, 20))

    if not rows:
        return {"error": "No price data found in Rates sheet"}

    # Deduplicate — compact normalization to catch M.S./MS, quotes, space variants
    def _compact(s):
        return re.sub(r"[^A-Z0-9]", "", str(s).upper())

    seen = {}  # compact_key -> row
    for r in rows:
        key = _compact(r[0])
        if key in seen:
            # Keep the one with the cleaner name (shorter or no special chars)
            existing_name = seen[key][0]
            if len(r[0]) <= len(existing_name):
                seen[key] = r
        else:
            seen[key] = r
    rows = list(seen.values())

    conn.execute("""
        CREATE TABLE IF NOT EXISTS component_price_master (
            item TEXT PRIMARY KEY, category TEXT,
            unit TEXT, price REAL, previous_price REAL,
            excel_row INTEGER, excel_col INTEGER
        )""")
    # Add columns if they don't exist (migration)
    for col in ['excel_row', 'excel_col']:
        try:
            conn.execute(f"ALTER TABLE component_price_master ADD COLUMN {col} INTEGER")
        except Exception:
            pass

    # Build compact->existing item map from DB to avoid inserting near-dupes
    existing = {}
    for (db_item,) in conn.execute("SELECT item FROM component_price_master"):
        existing[_compact(db_item)] = db_item

    for item, category, unit, price, prev, excel_row, excel_col in rows:
        ck = _compact(item)
        # If a near-duplicate already exists in DB with a different exact name, use that name
        if ck in existing and existing[ck] != item:
            item = existing[ck]
        conn.execute("""
            INSERT INTO component_price_master (item, category, unit, price, previous_price, excel_row, excel_col)
            VALUES (?,?,?,?,?,?,?)
            ON CONFLICT(item) DO UPDATE SET
                category=excluded.category, unit=excluded.unit,
                price=excluded.price, previous_price=excluded.previous_price,
                excel_row=excluded.excel_row, excel_col=excluded.excel_col
        """, (item, category, unit, price, prev, excel_row, excel_col))
        existing[ck] = item

    return {"rows": len(rows)}


# ─────────────────────────────────────────────────────────────────
# 2. HPU → hpu_master
# ─────────────────────────────────────────────────────────────────

def parse_hpu(xl, conn):
    sheet = _find_sheet(xl, "hpu")
    if sheet is None:
        return {"skipped": "HPU sheet not found"}

    df = xl.parse(sheet, header=None)
    row0 = df.iloc[0]  # "Costing of H..." titles (one per KW group)
    row1 = df.iloc[1]  # variant names: Duplex 1, Duplex 2, Simplex
    row2 = df.iloc[2]  # Duplex 1 col headers: S No., Items, Qty., Unit, Rate, Amount
    row3 = df.iloc[3]  # Duplex 2/Simplex col headers: Qty., Unit, Rate, Amount (row below Duplex 1 headers)

    col_map = {}
    title_cols = [(i, str(v)) for i, v in enumerate(row0) if pd.notna(v) and "Costing" in str(v)]
    variant_cols = [(i, str(v)) for i, v in enumerate(row1)
                    if pd.notna(v) and str(v).strip() not in ("nan", "")]

    for title_col, title_text in title_cols:
        kw_match = re.search(r'(\d+)\s*KW', title_text, re.I)
        kw = int(kw_match.group(1)) if kw_match else None

        next_titles = [tc for tc, _ in title_cols if tc > title_col]
        end = next_titles[0] if next_titles else df.shape[1]

        for var_col, var_name in variant_cols:
            if title_col <= var_col < end:
                for dc in range(var_col, min(var_col + 8, df.shape[1])):
                    # Duplex 1 headers are in row2; Duplex 2/Simplex headers are in row3
                    cell2 = str(row2.iloc[dc]).strip().lower() if pd.notna(row2.iloc[dc]) else ""
                    cell3 = str(row3.iloc[dc]).strip().lower() if pd.notna(row3.iloc[dc]) else ""
                    cell = cell2 or cell3
                    if "items" in cell or ("item" in cell and "s no" not in cell):
                        col_map[dc] = (kw, var_name, "item_col")
                    elif "qty" in cell:
                        col_map[dc] = (kw, var_name, "qty_col")
                    elif "unit" in cell:
                        col_map[dc] = (kw, var_name, "unit_col")
                    elif "rate" in cell:
                        col_map[dc] = (kw, var_name, "rate_col")
                    elif "amount" in cell:
                        col_map[dc] = (kw, var_name, "amount_col")

    # Duplex 2 and Simplex share the Items column with Duplex 1 for the same KW.
    # Build a map: kw -> item column index (from Duplex 1's item_col)
    item_col_by_kw = {kw: col_idx
                      for col_idx, (kw, _variant, field) in col_map.items()
                      if field == "item_col"}

    _skip_items = {"nan", "", "items", "total amount", "s no", "s no.", "s.no", "s.no."}

    records = []
    # Data starts at row index 4 (rows 0-3 are title + variant + two header rows)
    for _, row in df.iloc[4:].iterrows():
        block_data = {}
        for col_idx, (kw, variant, field) in col_map.items():
            key = (kw, variant)
            if key not in block_data:
                block_data[key] = {}
            val = row.iloc[col_idx] if col_idx < len(row) else None
            if pd.notna(val):
                block_data[key][field] = str(val).strip()

        for (kw, variant), fields in block_data.items():
            item = fields.get("item_col", "").strip()
            # Duplex 2 / Simplex have no item_col — borrow from Duplex 1's column
            if not item:
                ic = item_col_by_kw.get(kw)
                if ic is not None and ic < len(row):
                    raw = row.iloc[ic]
                    item = str(raw).strip() if pd.notna(raw) else ""
            if not item or item.lower() in _skip_items:
                continue
            qty    = safe_float(fields.get("qty_col"))
            rate   = safe_float(fields.get("rate_col"))
            # Compute live; fall back to cached amount (e.g. LABOUR CHARGES has amount but no qty/rate)
            amount = round(qty * rate, 2) if (qty and rate) else safe_float(fields.get("amount_col"))
            records.append({
                "unit_kw": kw,
                "variant": variant,
                "item": item.upper(),
                "qty":    qty,
                "unit":   fields.get("unit_col", "").strip() or None,
                "rate":   rate,
                "amount": amount,
            })

    df_out = pd.DataFrame(records).dropna(subset=["item"])
    df_out = df_out[df_out["item"].str.strip() != ""]
    df_out.to_sql("hpu_master", conn, if_exists="replace", index=False)
    return {"rows": len(df_out)}


# ─────────────────────────────────────────────────────────────────
# 2b. pumping_unit_price  (derived from hpu_master, excl. heater/thermostat)
# ─────────────────────────────────────────────────────────────────

PUMPING_UNIT_MARGIN = 1.8

def rebuild_pumping_unit_price(conn):
    """
    Repopulate pumping_unit_price from hpu_master.

    cost       = SUM(hpu_master.amount) per (unit_kw, variant), excluding
                 rows whose item name contains 'heater' or 'thermostat'
                 (a pumping unit is an HPU without the heating element).
    sell_price = cost × 1.8 (matches the HPU sheet's Excel markup).
    """
    c = conn.cursor()
    c.execute("DELETE FROM pumping_unit_price")
    rows = c.execute("""
        SELECT unit_kw, variant, ROUND(SUM(amount), 2) AS cost
        FROM hpu_master
        WHERE LOWER(item) NOT LIKE '%heater%'
          AND LOWER(item) NOT LIKE '%thermostat%'
        GROUP BY unit_kw, variant
    """).fetchall()
    c.executemany(
        "INSERT INTO pumping_unit_price (unit_kw, variant, cost, sell_price, margin_factor) "
        "VALUES (?, ?, ?, ?, ?)",
        [(kw, variant, cost, round(cost * PUMPING_UNIT_MARGIN, 2), PUMPING_UNIT_MARGIN)
         for (kw, variant, cost) in rows],
    )
    return {"rows": len(rows)}


# ─────────────────────────────────────────────────────────────────
# 3. BURNER → burner_pricelist_master
# ─────────────────────────────────────────────────────────────────

def parse_burner(xl, conn):
    """
    Compute burner pricelist from live Rates K/C column values.
    Replaces the stale cached-value reader (data_only=True returned wrong prices
    whenever K5/K6/etc. changed in the Rates sheet).

    Formula chain (verified against Pricelist WorkBook 28-08-2025):
      Oil Film BURNER ALONE = SUM(OilBurner group rows 4-8) * 2.5
      Gas      BURNER ALONE = Oil Film BURNER ALONE  (same size)
      HV Oil   BURNER ALONE = Oil Film BURNER ALONE  (same size)
    """
    import openpyxl
    import re as _re

    # ── 1. Read live Rates K col (rows 5-30) and C col (raw materials) ──
    wb = openpyxl.load_workbook(xl.io, read_only=True, data_only=True)
    _rs = next((s for s in wb.sheetnames if s.strip().lower() == "rates"), None)
    if not _rs:
        wb.close()
        return {"error": "Rates sheet not found"}
    ws_r = wb[_rs]
    k = {r: float(ws_r.cell(r, 11).value)
         for r in range(5, 31) if ws_r.cell(r, 11).value is not None}
    c_col = {}
    for _r in range(1, 40):
        _v = ws_r.cell(_r, 3).value
        try:
            if _v is not None:
                c_col[_r] = float(_v)
        except (ValueError, TypeError):
            pass
    wb.close()

    # ── 2. Open Oil Burner and HV Oil Burner sheets ──────────────────────
    _ob_sn = next((s for s in xl.sheet_names
                   if "oil burner" in s.lower() and "hv" not in s.lower()), None)
    _hv_sn = next((s for s in xl.sheet_names
                   if "oil burner" in s.lower() and "hv" in s.lower()), None)
    wb_f = openpyxl.load_workbook(xl.io, read_only=True, data_only=False)
    wb_v = openpyxl.load_workbook(xl.io, read_only=True, data_only=True)
    ws_ob_f = wb_f[_ob_sn] if _ob_sn else None
    ws_ob_v = wb_v[_ob_sn] if _ob_sn else None
    ws_hv_f = wb_f[_hv_sn] if _hv_sn else None
    ws_hv_v = wb_v[_hv_sn] if _hv_sn else None

    def _resolve(ws_f, row, col, depth=0):
        """Resolve a rate-cell formula to float, following cross-sheet refs."""
        if depth > 8 or ws_f is None:
            return 0.0
        v = ws_f.cell(row, col).value
        if v is None:
            return 0.0
        if isinstance(v, (int, float)):
            return float(v)
        if not isinstance(v, str):
            return 0.0
        # =Rates!K{n}  or  =Rates!$K${n}
        m = _re.search(r"Rates!.*?K\$?(\d+)", v, _re.I)
        if m:
            return k.get(int(m.group(1)), 0.0)
        # =Rates!C{n}
        m = _re.search(r"Rates!.*?C\$?(\d+)", v, _re.I)
        if m:
            return c_col.get(int(m.group(1)), 0.0)
        # =' Oil Burner'!{col}{row}  (cross-sheet ref from HV sheet)
        m = _re.match(r"^=.*?Oil Burner.*?!([A-Za-z]+)(\d+)$", v.strip(), _re.I)
        if m:
            rc = openpyxl.utils.column_index_from_string(m.group(1))
            rr = int(m.group(2))
            return _resolve(ws_ob_f, rr, rc, depth + 1)
        # ={col}{row}  (same-sheet cross-ref, e.g. =E7 or =W17)
        m = _re.match(r"^=([A-Za-z]+)(\d+)$", v.strip(), _re.I)
        if m:
            rc = openpyxl.utils.column_index_from_string(m.group(1))
            rr = int(m.group(2))
            return _resolve(ws_f, rr, rc, depth + 1)
        return 0.0

    def _tot(ws_f, ws_v, row, qty_col, rate_col, mc_col):
        """Compute TOTAL = qty x rate + mc for one Oil Burner row."""
        if ws_v is None:
            return 0.0
        qty_v = ws_v.cell(row, qty_col).value
        mc_v  = ws_v.cell(row, mc_col).value
        qty   = float(qty_v) if qty_v is not None else 0.0
        mc    = float(mc_v)  if mc_v  is not None else 0.0
        rate  = _resolve(ws_f, row, rate_col) if ws_f else 0.0
        return qty * rate + mc

    def _sum(ws_f, ws_v, rows, qty_col, rate_col, mc_col):
        return sum(_tot(ws_f, ws_v, r, qty_col, rate_col, mc_col) for r in rows)

    # Group column layouts: (qty_col, rate_col, mc_col)
    G2 = (3,  5,  7)   # Oil Burner 2A/3A
    G4 = (12, 14, 16)  # Oil Burner 4A
    G5 = (21, 23, 25)  # Oil Burner 5A/6A
    G7 = (30, 32, 34)  # Oil Burner 7A

    def _ob(row, g):  return _tot(ws_ob_f, ws_ob_v, row, *g)
    def _hv(row, g):  return _tot(ws_hv_f, ws_hv_v, row, *g)

    # ── 3. Compute Oil Film BURNER ALONE values ───────────────────────────
    ba2 = _sum(ws_ob_f, ws_ob_v, range(4, 9),  *G2) * 2.5
    ba3 = 14000.0                                         # hardcoded in Excel
    ba4 = _sum(ws_ob_f, ws_ob_v, range(4, 9),  *G4) * 2.5
    ba5 = _sum(ws_ob_f, ws_ob_v, range(4, 9),  *G5) * 2.5
    ba6 = ba5
    ba7 = _sum(ws_ob_f, ws_ob_v, range(4, 14), *G7) * 2.5

    # ── 4. Compute other Oil Film components ─────────────────────────────
    ys       = k.get(24, 450) * 2.5                       # Y-Strainer (all sizes)
    mv_base  = _ob(13, G4) * 2                            # Micro Valve 2A-6A (uses 4A col)
    mv7      = k.get(13, 2500) * 2                        # Micro Valve 7A

    bfv2 = _ob(9, G2) * 2;   bfv4 = _ob(9, G4) * 2
    bfv5 = _ob(9, G5) * 2;   bfv7 = _ob(17, G7) * 2

    fl2  = (_ob(11, G2) + _ob(12, G2)) * 2               # Flex hoses 2A/3A/4A
    fl5  = (_ob(11, G5) + _ob(12, G5)) * 2               # Flex hoses 5A/6A
    fl7  = (_ob(15, G7) + _ob(16, G7)) * 2               # Flex hoses 7A

    cp2  = _ob(16, G2) * 1.8;  cp3 = 6518.0              # CI Plate (3A hardcoded)
    cp4  = _ob(16, G4) * 1.8;  cp5 = _ob(16, G5) * 1.8
    cp6  = cp5;                 cp7 = _ob(20, G7) * 1.8

    bb2  = _ob(17, G2) * 2;    bb4 = _ob(17, G4) * 2
    bb5  = _ob(17, G5) * 2.2;  bb6 = bb5                 # 5A uses 2.2 markup
    bb7  = _ob(19, G7) * 2

    # ── 5. SG Assembly for Spares section ────────────────────────────────
    sg2 = _ob(8, G2) * 3
    sg3 = sg2 + 250
    sg4 = _ob(8, G4) * 3.93
    sg5 = _ob(8, G5) * 3.4
    sg6 = sg5 + 200
    sg7 = _ob(13, G7) * 2

    # ── 6. Ball Valves for Gas/Dual-Fuel sections ─────────────────────────
    gbv2 = round(2 * k.get(28, 1718) * 0.78, 2)
    gbv4 = round(2 * k.get(26, 2500) * 0.78, 2)
    gbv5 = round(2 * k.get(29, 2929) * 0.78, 2)
    gbv6 = round(2 * k.get(30, 5000) * 0.78, 2)
    gbv7 = gbv6

    # ── 7. HV flex hoses (from HV Oil Burner sheet) ───────────────────────
    hv_fl3 = _sum(ws_hv_f, ws_hv_v, range(7, 9),  *G2) * 2
    hv_fl4 = _sum(ws_hv_f, ws_hv_v, range(7, 9),  *G4) * 2
    hv_fl5 = _sum(ws_hv_f, ws_hv_v, range(7, 10), *G5) * 2   # 3 rows for 5A/6A
    hv_fl7 = _sum(ws_hv_f, ws_hv_v, range(14, 16),*G7) * 2

    wb_f.close()
    wb_v.close()

    # ── 8. Build records ──────────────────────────────────────────────────
    records = []

    def _r(section, size, comp, price):
        records.append({
            "section":     section,
            "burner_size": f"ENCON {size}",
            "component":   comp.upper(),
            "price":       round(float(price), 2),
        })

    # --- OIL FILM BURNER ---
    S_FILM = "PRICE FOR VARIOUS SIZES OF ENCON 'FILM' BURNER & ACCESSORIES"
    for sz, ba, bfv, cp, bb, fl, mv in [
        ("2A", ba2, bfv2, cp2, bb2, fl2, mv_base),
        ("3A", ba3, bfv2, cp3, bb2, fl2, mv_base),
        ("4A", ba4, bfv4, cp4, bb4, fl2, mv_base),
        ("5A", ba5, bfv5, cp5, bb5, fl5, mv_base),
        ("6A", ba6, bfv5, cp6, bb6, fl5, mv_base),
        ("7A", ba7, bfv7, cp7, bb7, fl7, mv7),
    ]:
        bs = ba + mv + cp + bb + fl + ys + bfv
        _r(S_FILM, sz, "BURNER ALONE", ba)
        _r(S_FILM, sz, "MICRO VALVE", mv)
        _r(S_FILM, sz, "C.I.BURNER PLATE", cp)
        _r(S_FILM, sz, "HIGH AL. WHYTEHEAT K BURNER BLOCK", bb)
        _r(S_FILM, sz, "FLEXIBLE HOSES SET", fl)
        _r(S_FILM, sz, "Y TYPE STRAINER", ys)
        _r(S_FILM, sz, "BUTTERFLY VALVE", bfv)
        _r(S_FILM, sz, "BURNER SET", bs)

    # --- SPARES ---
    S_SPARE = "PRICE LIST FOR SPARES OF IIP ENCON OIL FILM BURNERS"
    for sz, sg, ar in [
        ("ENCON-2A", sg2, 2900), ("ENCON-3A", sg3, 2900),
        ("ENCON-4A", sg4, 3600), ("ENCON-5A", sg5, 6600),
        ("ENCON-6A", sg6, 6600), ("ENCON-7A", sg7, 8100),
    ]:
        records.append({"section": S_SPARE, "burner_size": sz.upper(),
                        "component": "S.G. ASSEMBLY", "price": round(sg, 2)})
        records.append({"section": S_SPARE, "burner_size": sz.upper(),
                        "component": "AIR RESISTOR",  "price": ar})

    # --- DUAL FUEL ---
    S_DUAL = "PRICE FOR VARIOUS SIZES OF ENCON DUAL FUEL BURNER & ACCESSORIES"
    # Dual Fuel flex hoses: Oil Film flex * 3/2  (minus 5 for small sizes)
    dfl2 = fl2 * 1.5 - 5;  dfl5 = fl5 * 1.5;  dfl7 = fl7 * 1.5
    for sz, dba, dmv, dcp, dbb, dfl, dbv, dbfv in [
        ("2A", ba2+sg2, mv_base, cp2, bb2,    dfl2, gbv2, bfv2),
        ("3A", ba3+sg3, mv_base, cp3, bb2,    dfl2, gbv2, bfv2),
        ("4A", ba4+sg4, mv_base, cp4, bb4,    dfl2, gbv4, bfv4),
        ("5A", ba5+sg5, mv_base, cp5, 11850., dfl5, gbv5, bfv5),  # 11850 hardcoded
        ("6A", ba6+sg6, mv_base, cp6, bb6,    dfl5, gbv6, bfv5),
        ("7A", ba7+sg7, mv7,     cp7, bb7,    dfl7, gbv7, bfv7),
    ]:
        dbs = dba + dmv + dcp + dbb + dfl + dbv + ys + dbfv
        _r(S_DUAL, sz, "BURNER ALONE", dba)
        _r(S_DUAL, sz, "MICRO VALVE", dmv)
        _r(S_DUAL, sz, "C.I.BURNER PLATE", dcp)
        _r(S_DUAL, sz, "HIGH AL. WHYTEHEAT K BURNER BLOCK", dbb)
        _r(S_DUAL, sz, "FLEXIBLE HOSES SET", dfl)
        _r(S_DUAL, sz, "BALL VALVE", dbv)
        _r(S_DUAL, sz, "Y TYPE STRAINER", ys)
        _r(S_DUAL, sz, "BUTTERFLY VALVE", dbfv)
        _r(S_DUAL, sz, "BURNER SET", dbs)

    # --- HV OIL ---
    S_HV_OIL = "PRICE LIST FOR HIGH VELOCITY OIL BURNERS"
    for sz, hba, hbb, hmv, hfl, hbfv in [
        ("HV-3A", ba3, 44000.,  mv_base, hv_fl3, bfv2),
        ("HV-4A", ba4, 59000.,  mv_base, hv_fl4, bfv4),
        ("HV-5A", ba5, 71500.,  mv_base, hv_fl5, bfv5),
        ("HV-6A", ba6, 161000., mv_base, hv_fl5, bfv5),
        ("HV-7A", ba7, 201300., mv7,     hv_fl7, bfv7),
    ]:
        hbs = hba + hbb + hmv + hfl + ys + hbfv
        _r(S_HV_OIL, sz, "BURNER ALONE", hba)
        _r(S_HV_OIL, sz, "BURNER BLOCK", hbb)
        _r(S_HV_OIL, sz, "MICRO VALVE", hmv)
        _r(S_HV_OIL, sz, "FLEXIBLE HOSES SET", hfl)
        _r(S_HV_OIL, sz, "Y TYPE STRAINER", ys)
        _r(S_HV_OIL, sz, "BUTTERFLY VALVE", hbfv)
        _r(S_HV_OIL, sz, "BURNER SET", hbs)

    # --- HV GAS ---
    S_HV_GAS = "PRICE LIST FOR HIGH VELOCITY GAS BURNERS"
    for sz, gba, gbb, gbv, gfl, gbfv in [
        ("HV-3A", ba3, 44000.,  gbv2, hv_fl3, bfv2),
        ("HV-4A", ba4, 59000.,  gbv4, hv_fl4, bfv4),
        ("HV-5A", ba5, 71500.,  gbv5, hv_fl5, bfv5),
        ("HV-6A", ba6, 161000., gbv6, hv_fl5, bfv5),
        ("HV-7A", ba7, 201300., gbv7, hv_fl7, bfv7),
    ]:
        gbs = gba + gbb + gbv + gfl + gbfv
        _r(S_HV_GAS, sz, "BURNER ALONE", gba)
        _r(S_HV_GAS, sz, "BURNER BLOCK", gbb)
        _r(S_HV_GAS, sz, "BALL VALVE", gbv)
        _r(S_HV_GAS, sz, "FLEXIBLE HOSES SET", gfl)
        _r(S_HV_GAS, sz, "BUTTERFLY VALVE", gbfv)
        _r(S_HV_GAS, sz, "BURNER SET", gbs)

    # --- GAS (Film Gas) ---
    # BURNER ALONE = Oil Film BURNER ALONE (B54=B5 etc. in BURNER sheet)
    S_GAS = "PRICE FOR VARIOUS SIZES OF ENCON 'GAS' BURNER & ACCESSORIES"
    for sz, gba, gbv, gcp, gbb, gfl, gbfv in [
        ("2A", ba2, gbv2, cp2, bb2, fl2, bfv2),
        ("3A", ba3, gbv2, cp3, bb2, fl2, bfv2),
        ("4A", ba4, gbv4, cp4, bb4, fl2, bfv4),
        ("5A", ba5, gbv5, cp5, bb5, fl5, bfv5),
        ("6A", ba6, gbv6, cp6, bb6, fl5, bfv5),
        ("7A", ba7, gbv7, cp7, bb7, fl7, bfv7),
    ]:
        gbs = gba + gbv + gcp + gbb + gfl + gbfv
        _r(S_GAS, sz, "BURNER ALONE", gba)
        _r(S_GAS, sz, "BALL VALVE", gbv)
        _r(S_GAS, sz, "C.I.BURNER PLATE", gcp)
        _r(S_GAS, sz, "HIGH AL. WHYTEHEAT K BURNER BLOCK", gbb)
        _r(S_GAS, sz, "FLEXIBLE HOSES SET", gfl)
        _r(S_GAS, sz, "BUTTERFLY VALVE", gbfv)
        _r(S_GAS, sz, "BURNER SET", gbs)

    # --- PILOT BURNERS (hardcoded prices, no formula dependency) ---
    S_PILOT = "PRICE LIST FOR PILOT BURNER BURNERS"
    for model, price in [
        ("PB-LPG -10 KW", 11000), ("(NG/LPG) -100 KW", 21000),
        ("PB COG 100 KW", 32000), ("ELECTRODE FOR LPG PILOT", 1650),
        ("ELECTRODE FOR COG PILOT", 3300), ("BURNER CONTROL UNIT", 10000),
        ("IGNITION TRANSFORMER", 5500), ("UV SENSOR WITH AIR JACKET", 13000),
    ]:
        records.append({"section": S_PILOT,
                        "burner_size": f"ENCON {model}".upper(),
                        "component": "BURNER", "price": price})

    df_out = pd.DataFrame(records)
    df_out.to_sql("burner_pricelist_master", conn, if_exists="replace", index=False)
    return {"rows": len(df_out)}


# ─────────────────────────────────────────────────────────────────
# 4. BLOWER → blower_pricelist_master
# ─────────────────────────────────────────────────────────────────

def parse_blower(xl, conn):
    """
    Blower sheet layout (0-indexed cols):
      MEDIUM PRESSURE: section header row 1, data rows 3-13
      HIGH PRESSURE:   section header row 14, data rows 16-25

    Fixed column mapping for ENCON price rows:
      0=model, 1=hp, 2=cfm, 3=nm3_per_hr, 4=pressure,
      5=price_without_motor (ENCON selling price w/o motor),
      6=price_with_motor    (ENCON selling price w/ motor),
      7=blower_weight, 9=per_kg_amount, 10=motor_price_abb,
      11=perkin_price_without_motor, 12=perkin_price_with_motor

    Using fixed column indices avoids the misalignment caused by the
    old non-null stripping approach (which merged sparse columns).
    """
    sheet = _find_sheet(xl, "blower")
    if sheet is None:
        return {"skipped": "Blower sheet not found"}

    df = xl.parse(sheet, header=None)

    def cell(row_idx, col_idx):
        try:
            v = df.iloc[row_idx, col_idx]
            return None if pd.isna(v) else v
        except IndexError:
            return None

    def sval(row_idx, col_idx):
        v = cell(row_idx, col_idx)
        return str(v).strip() if v is not None else None

    def fval(row_idx, col_idx):
        return safe_float(cell(row_idx, col_idx))

    records = []

    # ── MEDIUM PRESSURE (28" W.G.) — rows 3-13 (0-indexed) ──
    for r in range(3, 14):
        model = sval(r, 0)
        if not model or not model.upper().startswith("ENCON"):
            continue
        records.append({
            "section":               "MEDIUM PRESSURE",
            "model":                 model,
            "hp":                    fval(r, 1),
            "cfm":                   fval(r, 2),
            "nm3_per_hr":            fval(r, 3),
            "pressure":              sval(r, 4),
            "price_without_motor":   fval(r, 5),
            "price_with_motor":      fval(r, 6),
            "blower_weight":         fval(r, 7),
            "per_kg_amount":         fval(r, 9),
            "motor_price_abb":       fval(r, 10),
            "perkin_price_wo_motor": fval(r, 11),
            "perkin_price_w_motor":  fval(r, 12),
        })

    # ── HIGH PRESSURE (40" W.G.) — rows 16-25 (0-indexed) ──
    for r in range(16, 26):
        model = sval(r, 0)
        if not model or not model.upper().startswith("ENCON"):
            continue
        records.append({
            "section":               "HIGH PRESSURE",
            "model":                 model,
            "hp":                    fval(r, 1),
            "cfm":                   fval(r, 2),
            "nm3_per_hr":            fval(r, 3),
            "pressure":              sval(r, 4),
            "price_without_motor":   fval(r, 5),
            "price_with_motor":      fval(r, 6),
            "blower_weight":         fval(r, 7),
            "per_kg_amount":         fval(r, 9),
            "motor_price_abb":       fval(r, 10),
            "perkin_price_wo_motor": fval(r, 11),
            "perkin_price_w_motor":  fval(r, 12),
        })

    # ── BLOWER DM 40 (direct-mount, 40" WG) — rows 38-44, model col 0, total col 8 ──
    for r in range(38, 45):
        model = sval(r, 0)
        if not model or not re.match(r'^\d', model):
            continue
        records.append({
            "section":             "BLOWER DM 40",
            "model":               model,
            "hp":                  None,
            "cfm":                 None,
            "nm3_per_hr":          None,
            "pressure":            None,
            "price_without_motor": None,
            "price_with_motor":    None,
            "blower_weight":       None,
            "per_kg_amount":       None,
            "motor_price_abb":     None,
            "perkin_price_wo_motor": None,
            "perkin_price_w_motor":  None,
            "total_cost":          fval(r, 8),
        })

    # ── BLOWER DM 28 (direct-mount, 28" WG) — rows 38-44, model col 11, total col 19 ──
    for r in range(38, 45):
        model = sval(r, 11)
        if not model or not re.match(r'^\d', model):
            continue
        records.append({
            "section":             "BLOWER DM 28",
            "model":               model,
            "hp":                  None,
            "cfm":                 None,
            "nm3_per_hr":          None,
            "pressure":            None,
            "price_without_motor": None,
            "price_with_motor":    None,
            "blower_weight":       None,
            "per_kg_amount":       None,
            "motor_price_abb":     None,
            "perkin_price_wo_motor": None,
            "perkin_price_w_motor":  None,
            "total_cost":          fval(r, 19),
        })

    # ── BLOWER IDM — rows 58-62, model col 0, total col 12 ──
    for r in range(58, 63):
        model = sval(r, 0)
        if not model or not re.match(r'^\d', model):
            continue
        records.append({
            "section":             "BLOWER IDM",
            "model":               model,
            "hp":                  None,
            "cfm":                 None,
            "nm3_per_hr":          None,
            "pressure":            None,
            "price_without_motor": None,
            "price_with_motor":    None,
            "blower_weight":       None,
            "per_kg_amount":       None,
            "motor_price_abb":     None,
            "perkin_price_wo_motor": None,
            "perkin_price_w_motor":  None,
            "total_cost":          fval(r, 12),
        })

    df_out = pd.DataFrame(records)
    df_out.to_sql("blower_pricelist_master", conn, if_exists="replace", index=False)

    # Also sync to blower_master (used by blower_selector.py)
    _sync_blower_master(df_out, conn)

    # ── DM / IDM component breakdown tables ─────────────────────────
    # Structure: one row per model with qty + cost columns for each material
    # DM 40 qty:  rows 38-44, cols 0(model),1-7(qty),8(total)
    # DM 40 cost: rows 47-54, cols 0(model),1-7(cost),8(subtotal),9(×1.3),10(×1.8=selling)
    # DM 28 qty:  rows 38-44, cols 11(model),12-18(qty),19(total)
    # DM 28 cost: rows 47-54, cols 11(model),12-18(cost),19(subtotal)  [no factor cols shown]
    # IDM qty:    rows 57-62, cols 0(model),1-11(qty),12(total)
    # IDM cost:   rows 65-70, cols 0(model),1-11(cost),12(subtotal),13(×1.3),14(×1.8=selling)

    DM_COMPS  = ["angle65_50","sheet8mm","sheet4mm","sheet2mm","flat","ci_hub","hardware"]
    IDM_COMPS = ["angle65_50","sheet8mm","sheet4mm","sheet2mm","flat","ci_hub","hardware","ms_round","plumber_block","coupling","channel"]

    dm_records = []

    # DM 40: qty rows 38-44 (0-idx), cost rows 47-54 (0-idx)
    dm40_qty  = {sval(r,0): [fval(r,c) for c in range(1,8)] for r in range(38,45) if sval(r,0) and re.match(r'^\d', sval(r,0))}
    dm40_cost = {sval(r,0): [fval(r,c) for c in range(1,8)] + [fval(r,8), fval(r,9), fval(r,10)] for r in range(47,55) if sval(r,0) and re.match(r'^\d', sval(r,0))}
    for model in sorted(dm40_qty):
        qtys  = dm40_qty.get(model, [None]*7)
        costs = dm40_cost.get(model, [None]*10)
        row = {"section":"BLOWER DM 40","model":model}
        for i,k in enumerate(DM_COMPS):
            row[k+"_qty"]  = qtys[i]  if i < len(qtys)  else None
            row[k+"_cost"] = costs[i] if i < len(costs) else None
        row["subtotal"]    = costs[7]  if len(costs)>7  else None
        row["factor_03"]   = costs[8]  if len(costs)>8  else None
        row["selling_price"]= costs[9] if len(costs)>9  else fval(list(dm40_qty.keys()).index(model)+38, 8)
        for k in IDM_COMPS[len(DM_COMPS):]:
            row[k+"_qty"] = None; row[k+"_cost"] = None
        dm_records.append(row)

    # DM 28: qty rows 38-44 (0-idx) cols 11-19, cost rows 47-54 (0-idx) cols 11-19
    dm28_qty  = {sval(r,11): [fval(r,c) for c in range(12,19)] for r in range(38,45) if sval(r,11) and re.match(r'^\d', sval(r,11))}
    dm28_cost = {sval(r,11): [fval(r,c) for c in range(12,19)] + [fval(r,19)] for r in range(47,55) if sval(r,11) and re.match(r'^\d', sval(r,11))}
    for model in sorted(dm28_qty):
        qtys  = dm28_qty.get(model, [None]*7)
        costs = dm28_cost.get(model, [None]*8)
        row = {"section":"BLOWER DM 28","model":model}
        for i,k in enumerate(DM_COMPS):
            row[k+"_qty"]  = qtys[i]  if i < len(qtys)  else None
            row[k+"_cost"] = costs[i] if i < len(costs) else None
        row["subtotal"]     = costs[7] if len(costs)>7 else None
        sub = costs[7] if len(costs)>7 else 0
        row["factor_03"]    = round(sub * 1.3, 2) if sub else None
        row["selling_price"]= round(sub * 1.3 * 1.8, 2) if sub else None
        for k in IDM_COMPS[len(DM_COMPS):]:
            row[k+"_qty"] = None; row[k+"_cost"] = None
        dm_records.append(row)

    # IDM: qty rows 57-62, cost rows 65-70
    idm_qty  = {sval(r,0): [fval(r,c) for c in range(1,12)] for r in range(57,63) if sval(r,0) and re.match(r'^\d', sval(r,0))}
    idm_cost = {sval(r,0): [fval(r,c) for c in range(1,12)] + [fval(r,12), fval(r,13), fval(r,14)] for r in range(65,71) if sval(r,0) and re.match(r'^\d', sval(r,0))}
    for model in sorted(idm_qty):
        qtys  = idm_qty.get(model, [None]*11)
        costs = idm_cost.get(model, [None]*14)
        row = {"section":"BLOWER IDM","model":model}
        for i,k in enumerate(IDM_COMPS):
            row[k+"_qty"]  = qtys[i]  if i < len(qtys)  else None
            row[k+"_cost"] = costs[i] if i < len(costs) else None
        row["subtotal"]     = costs[11] if len(costs)>11 else None
        row["factor_03"]    = costs[12] if len(costs)>12 else None
        row["selling_price"]= costs[13] if len(costs)>13 else None
        dm_records.append(row)

    if dm_records:
        pd.DataFrame(dm_records).to_sql("blower_dm_idm_master", conn, if_exists="replace", index=False)

    return {"rows": len(df_out)}


def _sync_blower_master(df_pricelist: pd.DataFrame, conn):
    """
    Normalise blower_pricelist_master into the fixed-column schema that
    blower_selector.py expects:
        model, hp, airflow, cfm, pressure, price_basic, price_premium

    ENCON prices (cols F/G): read directly — manually set selling prices.
    PERKIN prices (cols L/M): computed from input amounts (col J, K):
        price_without_motor = J × 1.8
        price_with_motor    = K × 1.5 + J × 1.8
    This replicates the Excel formulas =J*1.8 and =K*1.5+L so that if
    J or K changes in the sheet, Python recalculates without needing
    Excel to recalculate first.
    """
    if df_pricelist.empty:
        return

    cols = df_pricelist.columns.tolist()

    def _find_col(*keywords):
        for c in cols:
            cl = c.lower()
            if all(kw in cl for kw in keywords):
                return c
        return None

    airflow_col  = _find_col("nm3") or _find_col("airflow") or _find_col("flow")
    cfm_col      = _find_col("cfm")
    pressure_col = _find_col("pressure")

    # ENCON price columns (manually set, read directly)
    encon_basic_col = _find_col("price_without") or _find_col("price without")
    # "price with motor" but NOT "without"
    encon_prem_col = None
    for c in cols:
        cl = c.lower()
        if "price" in cl and "with" in cl and "motor" in cl and "without" not in cl:
            encon_prem_col = c
            break

    # PERKIN input columns: Amount (J) and Motor Price (K)
    amount_col      = _find_col("amount")
    motor_price_col = _find_col("motor_price") or _find_col("motor price")

    rows = []
    for _, r in df_pricelist.iterrows():
        model = r.get("model")
        if not model or str(model).lower() in ("nan", ""):
            continue

        # Try ENCON price first; fall back to computing from PERKIN amounts
        price_basic = safe_float(r[encon_basic_col]) if encon_basic_col else None
        price_prem  = safe_float(r[encon_prem_col])  if encon_prem_col  else None

        # If ENCON prices are missing/zero, compute from PERKIN amounts
        if not price_basic and amount_col:
            j = safe_float(r[amount_col]) or 0
            k = safe_float(r[motor_price_col]) if motor_price_col else 0
            if j:
                price_basic = round(j * 1.8, 2)
                price_prem  = round(k * 1.5 + price_basic, 2) if k else None

        rows.append((
            str(model),
            str(r.get("hp", "")) if "hp" in cols else None,
            str(r[airflow_col])  if airflow_col  else None,
            str(r[cfm_col])      if cfm_col      else None,
            str(r[pressure_col]) if pressure_col else None,
            str(price_basic)     if price_basic is not None else None,
            str(price_prem)      if price_prem  is not None else None,
        ))

    conn.execute("DELETE FROM blower_master")
    conn.executemany(
        "INSERT INTO blower_master "
        "(model, hp, airflow, cfm, pressure, price_basic, price_premium) "
        "VALUES (?,?,?,?,?,?,?)",
        rows,
    )


# ─────────────────────────────────────────────────────────────────
# 5 & 6. HORIZONTAL / VERTICAL → horizontal_master / vertical_master
#
# Amounts are computed live from openpyxl formulas:
#   MS Structure  : qty_kg × Rates!C{n} × 2.1  (C12 or C13 = plate rate)
#   Ceramic Fiber : 2.1 × rolls × Rates!C29    (C29 = roll price)
#   Panel/Pipeline/Trolley: =2.1*{constant} or =1.8*{constant} (computed directly)
#   Combustion sub-rows (BURNER!, Blower!, HPU!): cached stale value is fine —
#     ladle_params.py reads the HPU kW from the particular-name string, not the amount.
# ─────────────────────────────────────────────────────────────────

def _parse_ladle_master(xl, conn, sheet_keyword, table_name, blocks):
    """
    Parse a Horizontal or Vertical master sheet into SQLite.

    blocks: list of (header_row, data_rows_range, base_cols_list)
    Column layout per group: sno=base_col, particular=+1, qty=+2, amount=+3
    """
    import openpyxl
    import re as _re

    sheet = _find_sheet(xl, sheet_keyword)
    if sheet is None:
        return {"skipped": f"'{sheet_keyword}' sheet not found"}

    # ── Live Rates C column (C12/C13 = MS plate, C29 = ceramic fiber roll) ──
    wb_rates = openpyxl.load_workbook(xl.io, read_only=True, data_only=True)
    _rs = next((s for s in wb_rates.sheetnames if s.strip().lower() == "rates"), None)
    if not _rs:
        wb_rates.close()
        return {"error": "Rates sheet not found"}
    ws_r = wb_rates[_rs]
    c_col = {}
    for _r in range(1, 40):
        _v = ws_r.cell(_r, 3).value
        try:
            if _v is not None:
                c_col[_r] = float(_v)
        except (ValueError, TypeError):
            pass
    wb_rates.close()

    wb_f = openpyxl.load_workbook(xl.io, read_only=True, data_only=False)
    wb_v = openpyxl.load_workbook(xl.io, read_only=True, data_only=True)
    _sn = next((s for s in wb_f.sheetnames
                if sheet_keyword.lower() in s.strip().lower()), None)
    if _sn is None:
        wb_f.close(); wb_v.close()
        return {"skipped": f"'{sheet_keyword}' sheet not found"}
    ws_f = wb_f[_sn]
    ws_v = wb_v[_sn]

    def _compute_amount(row, amt_col):
        formula = ws_f.cell(row, amt_col).value
        cached  = ws_v.cell(row, amt_col).value

        if isinstance(formula, (int, float)):
            return float(formula)

        if isinstance(formula, str) and formula.startswith('='):
            # Rates!C{n} reference → substitute live value and evaluate
            m_r = _re.search(r'Rates!.*?C\$?(\d+)', formula, _re.I)
            if m_r:
                live_rate = c_col.get(int(m_r.group(1)), 0.0)
                if live_rate:
                    f_sub = _re.sub(r'Rates!.*?C\$?\d+', str(live_rate),
                                    formula, flags=_re.I).lstrip('=')
                    m_rnd = _re.match(r'ROUND\((.+),\s*(-?\d+)\)$', f_sub, _re.I)
                    if m_rnd:
                        inner, digits = m_rnd.group(1), int(m_rnd.group(2))
                        if _re.match(r'^[\d\.\*\+\-/\(\)\s]+$', inner):
                            try:
                                return round(eval(inner), digits)
                            except Exception:
                                pass
                    elif _re.match(r'^[\d\.\*\+\-/\(\)\s]+$', f_sub):
                        try:
                            return round(eval(f_sub), 2)
                        except Exception:
                            pass

            # No cross-sheet refs: pure constant formula (=2.1*60000, =1.8*69000, etc.)
            if '!' not in formula:
                f_clean = formula.lstrip('=')
                m_rnd = _re.match(r'ROUND\((.+),\s*(-?\d+)\)$', f_clean, _re.I)
                if m_rnd:
                    inner, digits = m_rnd.group(1), int(m_rnd.group(2))
                    if _re.match(r'^[\d\.\*\+\-/\(\)\s]+$', inner):
                        try:
                            return round(eval(inner), digits)
                        except Exception:
                            pass
                elif _re.match(r'^[\d\.\*\+\-/\(\)\s]+$', f_clean):
                    try:
                        return round(eval(f_clean), 2)
                    except Exception:
                        pass

        # Fallback: stale cached value (used for cross-sheet refs: BURNER!, Blower!, HPU!)
        try:
            return float(cached) if cached is not None else None
        except (TypeError, ValueError):
            return None

    records = []
    for header_row, data_rows, base_cols in blocks:
        for base_col in base_cols:
            name_col = base_col + 1
            qty_col  = base_col + 2
            amt_col  = base_col + 3

            model_v = (ws_v.cell(header_row, name_col).value or
                       ws_f.cell(header_row, name_col).value)
            model_name = str(model_v or '').strip().upper()
            # Strip "DESCRIPTION" column-header prefix that Excel sometimes
            # embeds in the same cell as the model subtitle e.g.
            # "DESCRIPTION( VERTICAL 10 TON LPS)" → "VERTICAL 10 TON LPS"
            if model_name.upper().startswith('DESCRIPTION'):
                model_name = model_name[len('DESCRIPTION'):].strip('( )\t')
            if not model_name:
                continue

            for r in data_rows:
                part_v = (ws_v.cell(r, name_col).value or
                          ws_f.cell(r, name_col).value)
                particular = str(part_v or '').strip().upper()
                if not particular or particular in (
                        'TOTAL', 'QTY', 'AMOUNT(RS.)', 'S.NO.', ''):
                    continue
                if 'COMBUSTION EQUIPMENT' in particular:
                    continue  # section sub-header, no amount

                qty_v   = ws_v.cell(r, qty_col).value
                qty_str = str(qty_v).strip() if qty_v is not None else None

                amount = _compute_amount(r, amt_col)
                records.append({
                    "model":      model_name,
                    "particular": particular,
                    "qty":        qty_str,
                    "amount":     amount,
                })

    wb_f.close()
    wb_v.close()

    df_out = pd.DataFrame(records)
    if not df_out.empty:
        df_out = df_out[["model", "particular", "qty", "amount"]]
    df_out.to_sql(table_name, conn, if_exists="replace", index=False)
    return {"rows": len(df_out)}


def parse_horizontal(xl, conn):
    # 3 groups per block at cols 1, 6, 11
    # Block 1: header row 1, data rows 2-10  (10T, 15T, 20T)
    # Block 2: header row 13, data rows 14-22 (30T, 35-40T, 60T)
    return _parse_ladle_master(xl, conn, "horizontal", "horizontal_master", [
        (1,  range(2, 11),  [1, 6, 11]),
        (13, range(14, 23), [1, 6, 11]),
    ])


# ─────────────────────────────────────────────────────────────────
# 6. VERTICAL → vertical_master
# ─────────────────────────────────────────────────────────────────

def parse_vertical(xl, conn):
    # Block 1: header row 2, data rows 3-10, groups at cols 1, 6, 11
    # Block 2: header row 13, data rows 14-21, groups at cols 1, 6, 11
    # Block 3: header row 24, data rows 25-32, group at col 1 only
    return _parse_ladle_master(xl, conn, "vertical", "vertical_master", [
        (2,  range(3, 11),  [1, 6, 11]),
        (13, range(14, 22), [1, 6, 11]),
        (24, range(25, 33), [1]),
    ])


# ─────────────────────────────────────────────────────────────────
# 7. PARTS SHEETS (Oil Burner, HV Oil Burner, Gas Burner)
# ─────────────────────────────────────────────────────────────────

def _parse_parts_sheet(xl, sheet_name, table_name, conn):
    sheet = _find_sheet(xl, sheet_name)
    if sheet is None:
        return {"skipped": f"'{sheet_name}' sheet not found"}

    df = xl.parse(sheet, header=None).dropna(how="all")
    records = []
    current_section = None

    for _, row in df.iterrows():
        vals = [str(x).strip() for x in row if pd.notna(x) and str(x).strip() not in ("nan", "")]
        if not vals:
            continue
        if len(vals) == 1 and safe_float(vals[0]) is None:
            current_section = vals[0].upper()
            continue
        if len(vals) < 2:
            continue
        if vals[0].upper() in ("S.NO.", "SNO", "SL NO"):
            continue

        idx = 1 if re.match(r'^\d+$', vals[0]) else 0
        particular = vals[idx].upper().strip() if idx < len(vals) else None
        if not particular or particular in ("PARTICULARS", "TOTAL", ""):
            continue

        qty = unit = rate = amount = None
        rest = vals[idx + 1:]
        for v in rest:
            if v.upper() in ("KGS", "KG", "NOS", "NO", "SET", "MTR", "ROLLS"):
                unit = v.upper()
            else:
                f = safe_float(v)
                if f is not None:
                    if rate is None:
                        rate = f
                    else:
                        amount = f
        for v in rest:
            f = safe_float(v)
            if f is not None and qty is None and f != rate and f != amount:
                qty = f
                break

        records.append({"section": current_section, "particular": particular,
                        "qty": qty, "unit": unit, "rate": rate, "amount": amount})

    pd.DataFrame(records).to_sql(table_name, conn, if_exists="replace", index=False)
    return {"rows": len(records)}


# ─────────────────────────────────────────────────────────────────
# 7a. Multi-column parts sheet helper (Oil Burner, HV Oil Burner)
#
# Both sheets have 4 groups side-by-side; each group uses 9 columns:
#   base+0=SNo, base+1=Particular, base+2=QTY, base+3=Unit,
#   base+4=Rate(formula), base+5=Amount, base+6=MC, base+7=Total
#
# Rate cells contain formulas like =Rates!K5 or cross-refs like =E7.
# We resolve these against live Rates K/C values instead of using
# the stale cached values produced by data_only=True.
# ─────────────────────────────────────────────────────────────────

def _parse_multicolumn_parts(xl, sheet_name, table_name, groups, conn):
    """
    groups: list of (header_row, data_rows_range, base_col_1indexed)
    Writes section-header rows (amount=None) followed by data rows with
    live-computed amounts into table_name.
    """
    import openpyxl
    import re as _re

    sheet = _find_sheet(xl, sheet_name)
    if sheet is None:
        return {"skipped": f"'{sheet_name}' sheet not found"}

    # Live Rates K and C columns
    wb = openpyxl.load_workbook(xl.io, read_only=True, data_only=True)
    _rs = next((s for s in wb.sheetnames if s.strip().lower() == "rates"), None)
    if not _rs:
        wb.close()
        return {"error": "Rates sheet not found"}
    ws_r = wb[_rs]
    k = {r: float(ws_r.cell(r, 11).value)
         for r in range(5, 31) if ws_r.cell(r, 11).value is not None}
    c_col = {}
    for _ri in range(1, 40):
        _v = ws_r.cell(_ri, 3).value
        try:
            if _v is not None:
                c_col[_ri] = float(_v)
        except (ValueError, TypeError):
            pass
    wb.close()

    wb_f = openpyxl.load_workbook(xl.io, read_only=True, data_only=False)
    wb_v = openpyxl.load_workbook(xl.io, read_only=True, data_only=True)
    ws_f = wb_f[sheet]
    ws_v = wb_v[sheet]

    # Also open Oil Burner sheet for cross-sheet refs (used by HV Oil Burner 7A)
    _ob_sn = next((s for s in wb_f.sheetnames
                   if "oil burner" in s.lower() and "hv" not in s.lower()), None)
    ws_ob_f = wb_f[_ob_sn] if _ob_sn else None

    def _resolve(ws, row, col, depth=0):
        if depth > 8 or ws is None:
            return 0.0
        v = ws.cell(row, col).value
        if v is None:
            return 0.0
        if isinstance(v, (int, float)):
            return float(v)
        if not isinstance(v, str):
            return 0.0
        m = _re.search(r"Rates!.*?K\$?(\d+)", v, _re.I)
        if m:
            return k.get(int(m.group(1)), 0.0)
        m = _re.search(r"Rates!.*?C\$?(\d+)", v, _re.I)
        if m:
            return c_col.get(int(m.group(1)), 0.0)
        # Cross-sheet ref to Oil Burner: =' Oil Burner'!{col}{row}
        m = _re.match(r"^=.*?Oil Burner.*?!([A-Za-z]+)(\d+)$", v.strip(), _re.I)
        if m:
            rc = openpyxl.utils.column_index_from_string(m.group(1))
            rr = int(m.group(2))
            return _resolve(ws_ob_f, rr, rc, depth + 1)
        # Same-sheet cross-ref: ={col}{row}
        m = _re.match(r"^=([A-Za-z]+)(\d+)$", v.strip(), _re.I)
        if m:
            rc = openpyxl.utils.column_index_from_string(m.group(1))
            rr = int(m.group(2))
            return _resolve(ws, rr, rc, depth + 1)
        return 0.0

    def _row_total(row, base_col):
        total_col  = base_col + 7
        amount_col = base_col + 5
        mc_col     = base_col + 6

        # 1. Prefer cached total — reflects Excel's last-saved state, correct even
        #    when items include assembly/MC costs not derivable from qty×rate alone
        tv = ws_v.cell(row, total_col).value
        if tv is not None:
            try: return float(tv)
            except (ValueError, TypeError): pass

        # 2. Hardcoded literal in formula workbook (labour, paint, etc.)
        tf = ws_f.cell(row, total_col).value
        if isinstance(tf, (int, float)):
            return float(tf)

        # 3. Fall back to cached amount + cached mc
        av   = ws_v.cell(row, amount_col).value
        mc_v = ws_v.cell(row, mc_col).value
        mc   = float(mc_v) if mc_v is not None else 0.0
        if av is not None:
            try: return float(av) + mc
            except (ValueError, TypeError): pass

        return mc

    records = []
    for header_row, data_rows, base_col in groups:

        title_v = ws_v.cell(header_row, base_col).value
        if not title_v:
            continue
        records.append({
            "section":    None,
            "particular": str(title_v).strip(),
            "qty":        None,
            "unit":       None,
            "rate":       None,
            "amount":     None,
        })
        for row in data_rows:
            part_v = ws_v.cell(row, base_col + 1).value
            if not part_v:
                continue
            particular = str(part_v).strip().upper()
            if not particular or particular in ("PARTICULARS", "TOTAL"):
                continue
            qty_v  = ws_v.cell(row, base_col + 2).value
            unit_v = ws_v.cell(row, base_col + 3).value
            rate_v = ws_v.cell(row, base_col + 4).value   # cached rate
            rate   = float(rate_v) if rate_v is not None else None
            total  = _row_total(row, base_col)
            records.append({
                "section":    None,
                "particular": particular,
                "qty":        float(qty_v) if qty_v is not None else None,
                "unit":       str(unit_v).strip() if unit_v else None,
                "rate":       rate,
                "amount":     round(total, 2) if total else None,
            })

    wb_f.close()
    wb_v.close()
    pd.DataFrame(records).to_sql(table_name, conn, if_exists="replace", index=False)
    return {"rows": len(records)}


def parse_oil_burner_parts(xl, conn):
    """
    Oil Burner sheet: 4 groups side-by-side.
      2A/3A: base col 1,  data rows 4-19
      4A:    base col 10, data rows 4-19
      5A/6A: base col 19, data rows 4-19
      7A:    base col 28, data rows 4-23
    """
    return _parse_multicolumn_parts(xl, "oil burner", "oil_burner_parts_master", [
        (1, range(4, 20), 1),
        (1, range(4, 20), 10),
        (1, range(4, 20), 19),
        (1, range(4, 24), 28),
    ], conn)


def parse_hv_oil_burner_parts(xl, conn):
    """
    HV Oil Burner sheet: 4 groups side-by-side.
      2A/3A: base col 1,  data rows 3-13
      4A:    base col 10, data rows 3-13
      5A/6A: base col 19, data rows 3-14
      7A:    base col 28, data rows 3-21
    """
    return _parse_multicolumn_parts(xl, "hv  oil burner", "hv_oil_burner_parts_master", [
        (1, range(3, 14), 1),
        (1, range(3, 14), 10),
        (1, range(3, 15), 19),
        (1, range(3, 22), 28),
    ], conn)


# ─────────────────────────────────────────────────────────────────
# 7b. GAS BURNER PARTS → gas_burner_parts_master
#
# The Gas Burner sheet has 5 groups side-by-side (multi-column layout).
# The generic _parse_parts_sheet fails on this layout because it reads
# all non-null values from a full row regardless of group boundaries.
# This dedicated parser handles each group's fixed column offset and
# computes amounts from live Rates K/C values.
#
# Groups:
#   50 NM³ GAIL   : header row 1, data rows  4-13, base col 1
#   100 NM³ GAIL  : header row 1, data rows  4-13, base col 10
#   150 NM³ GAIL  : header row 1, data rows  4-13, base col 19
#   100 NM³ PowerTrade: header row 17, data rows 20-31, base col 1
#   100 NM³ Ada   : header row 17, data rows 20-31, base col 10
# ─────────────────────────────────────────────────────────────────

def parse_gas_burner_parts(xl, conn):
    import openpyxl
    import re as _re

    sheet = _find_sheet(xl, "gas burner")
    if sheet is None:
        return {"skipped": "Gas Burner sheet not found"}

    # Live Rates K and C columns
    wb = openpyxl.load_workbook(xl.io, read_only=True, data_only=True)
    _rs = next((s for s in wb.sheetnames if s.strip().lower() == "rates"), None)
    if not _rs:
        wb.close()
        return {"error": "Rates sheet not found"}
    ws_r = wb[_rs]
    k = {r: float(ws_r.cell(r, 11).value)
         for r in range(5, 31) if ws_r.cell(r, 11).value is not None}
    c_col = {}
    for _ri in range(1, 40):
        _v = ws_r.cell(_ri, 3).value
        try:
            if _v is not None:
                c_col[_ri] = float(_v)
        except (ValueError, TypeError):
            pass
    wb.close()

    wb_f = openpyxl.load_workbook(xl.io, read_only=True, data_only=False)
    wb_v = openpyxl.load_workbook(xl.io, read_only=True, data_only=True)
    ws_gb_f = wb_f[sheet]
    ws_gb_v = wb_v[sheet]

    def _resolve_gb(row, col, depth=0):
        if depth > 8:
            return 0.0
        v = ws_gb_f.cell(row, col).value
        if v is None:
            return 0.0
        if isinstance(v, (int, float)):
            return float(v)
        if not isinstance(v, str):
            return 0.0
        m = _re.search(r"Rates!.*?K\$?(\d+)", v, _re.I)
        if m:
            return k.get(int(m.group(1)), 0.0)
        m = _re.search(r"Rates!.*?C\$?(\d+)", v, _re.I)
        if m:
            return c_col.get(int(m.group(1)), 0.0)
        m = _re.match(r"^=([A-Za-z]+)(\d+)$", v.strip(), _re.I)
        if m:
            rc = openpyxl.utils.column_index_from_string(m.group(1))
            rr = int(m.group(2))
            return _resolve_gb(rr, rc, depth + 1)
        return 0.0

    def _row_total(row, base_col):
        """Return computed or hardcoded total for a data row."""
        total_col  = base_col + 7
        amount_col = base_col + 5
        mc_col     = base_col + 6

        # 1. Prefer cached total — reflects the Excel's last-saved state and is
        #    correct even when Rates formulas reference stale cross-sheet values
        tv = ws_gb_v.cell(row, total_col).value
        if tv is not None:
            try: return float(tv)
            except (ValueError, TypeError): pass

        # 2. Total cell is a hardcoded literal in the formula workbook (Labour, Paint)
        tf = ws_gb_f.cell(row, total_col).value
        if isinstance(tf, (int, float)):
            return float(tf)

        # 3. Fall back to cached amount + cached mc
        av   = ws_gb_v.cell(row, amount_col).value
        mc_v = ws_gb_v.cell(row, mc_col).value
        mc   = float(mc_v) if mc_v is not None else 0.0
        if av is not None:
            try: return float(av) + mc
            except (ValueError, TypeError): pass

        return mc

    records = []
    # (header_row, data_rows, base_col)
    groups = [
        (1,  range(4,  14), 1),
        (1,  range(4,  14), 10),
        (1,  range(4,  14), 19),
        (17, range(20, 32), 1),
        (17, range(20, 32), 10),
    ]
    for header_row, data_rows, base_col in groups:
        title_v = ws_gb_v.cell(header_row, base_col).value
        if not title_v:
            continue
        # Section header row — amount=None signals burner_fabrication.py
        records.append({
            "section":    None,
            "particular": str(title_v).strip(),
            "qty":        None,
            "unit":       None,
            "rate":       None,
            "amount":     None,
        })
        for row in data_rows:
            part_v = ws_gb_v.cell(row, base_col + 1).value
            if not part_v:
                continue
            particular = str(part_v).strip().upper()
            if not particular or particular in ("PARTICULARS", "TOTAL"):
                continue
            qty_v  = ws_gb_v.cell(row, base_col + 2).value
            unit_v = ws_gb_v.cell(row, base_col + 3).value
            # Use cached rate (Excel's saved value); avoids wrong Rates-formula resolution
            rate_v = ws_gb_v.cell(row, base_col + 4).value
            rate   = float(rate_v) if rate_v is not None else None
            total  = _row_total(row, base_col)
            records.append({
                "section":    None,
                "particular": particular,
                "qty":        float(qty_v) if qty_v is not None else None,
                "unit":       str(unit_v).strip() if unit_v else None,
                "rate":       rate,
                "amount":     round(total, 2) if total else None,
            })

    wb_f.close()
    wb_v.close()
    pd.DataFrame(records).to_sql("gas_burner_parts_master", conn,
                                 if_exists="replace", index=False)
    return {"rows": len(records)}


# ─────────────────────────────────────────────────────────────────
# 8. RECUPERATOR → recuperator_master
#
# Replicates the Excel formulas exactly:
#
#   Type F  : col_cost = (qty / 0.3) × rate  for tube_c
#                        qty × rate           for all others
#             selling  = SUM(cols B–G) × 1.3 × 1.8
#
#   Type HT : selling  = SUM(cols N–Q) × 1.3 × 1.8
#
#   Type FS : col_cost = qty × 4.27 × rate   for tube_c and tube_b
#                        qty × rate           for all others
#             selling  = SUM(cols B–I) × 1.3 × 1.8
#
# Material prices come from component_price_master (already populated
# by parse_rates which runs first in parse_all).
# ─────────────────────────────────────────────────────────────────

def parse_recuperator(xl, conn):
    sheet = _find_sheet(xl, "recuperator")
    if sheet is None:
        return {"skipped": "Recuperator sheet not found"}

    # ── fetch live rates from component_price_master ──────────────
    def _rate(item_prefix):
        try:
            row = conn.execute(
                "SELECT price FROM component_price_master WHERE item LIKE ?",
                (item_prefix.strip() + '%',)
            ).fetchone()
            return float(row[0]) if row else 0.0
        except Exception:
            return 0.0

    R = {
        'tube_c':    _rate('M.S. Tube "C" Class'),
        'tube_b':    _rate('M.S. Tube "B" Class'),
        'ci_gills':  _rate('C.I. Gills'),
        'ang_6550':  _rate('M.S. Angle 65,50'),
        'ang_100':   _rate('M.S. Angle 100,100'),
        'plate_5mm': _rate('M.S. Plate 16mm* 5mm'),
        'bolt':      _rate('Hardware Bolt'),
        'channel':   _rate('M.S. Chanel'),
        'plate_10mm':_rate('M.S. Plate 16mm*10mm'),
        'plate_5mm2':_rate('M.S. Plate 5mm'),
        'ang_50':    _rate('M.S. Angle 50*6'),
        'ss_sheet':  _rate('S.S. Sheet 3mm'),
    }

    # ── read the raw quantity table (data_only to get values not formulas) ──
    import openpyxl
    wb_vals = openpyxl.load_workbook(xl.io, read_only=True, data_only=True)
    ws = wb_vals[sheet]
    rows_vals = list(ws.iter_rows(values_only=True))
    wb_vals.close()

    def _v(row_idx, col_idx):
        """0-based row and col. Returns float or 0."""
        try:
            v = rows_vals[row_idx][col_idx]
            return float(v) if v is not None else 0.0
        except Exception:
            return 0.0

    MARKUP = 1.3 * 1.8  # = 2.34

    records = []

    # ── Type F (rows 3–8, 0-indexed 2–7) ─────────────────────────
    # cols: A=0,B=1,C=2,D=3,E=4,F=5,G=6,H=7
    F_MODELS_ROWS = range(2, 8)   # rows 3–8
    for ri in F_MODELS_ROWS:
        model = rows_vals[ri][0]
        if model is None or not str(model).strip():
            continue
        qtc = _v(ri, 1)   # MS Tube C qty (mtr)
        qcg = _v(ri, 2)   # CI Gills qty
        qa6 = _v(ri, 3)   # Angle 65,50 (kg)
        qa1 = _v(ri, 4)   # Angle 100,100 (kg)
        qpl = _v(ri, 5)   # Plate 16mm*5mm (kg)
        qbl = _v(ri, 6)   # Hardware Bolt (kg)
        qch = _v(ri, 7)   # MS Channel (kg) — stored but excluded from Excel SUM
        tube_cost   = (qtc / 0.3) * R['tube_c'] if qtc else 0
        ci_cost     = qcg * R['ci_gills']
        ang65_cost  = qa6 * R['ang_6550']
        ang100_cost = qa1 * R['ang_100']
        plate_cost  = qpl * R['plate_5mm']
        bolt_cost   = qbl * R['bolt']
        chan_cost   = qch * R['channel']
        fab_cost    = tube_cost + ci_cost + ang65_cost + ang100_cost + plate_cost + bolt_cost
        selling     = round(fab_cost * MARKUP, 2)
        records.append({
            "type": "F", "model": str(model).strip(),
            "ms_tube_c_qty": qtc, "ms_tube_c_cost": round(tube_cost, 2),
            "ci_gills_qty":  qcg, "ci_gills_cost":  round(ci_cost, 2),
            "ang_6550_qty":  qa6, "ang_6550_cost":  round(ang65_cost, 2),
            "ang_100100_qty":qa1, "ang_100100_cost": round(ang100_cost, 2),
            "plate_16_5_qty":qpl, "plate_16_5_cost": round(plate_cost, 2),
            "bolt_qty":      qbl, "bolt_cost":       round(bolt_cost, 2),
            "channel_qty":   qch, "channel_cost":    round(chan_cost, 2),
            "fabrication_cost": round(fab_cost, 2), "selling_price": selling,
        })

    # ── Type HT (same input rows, cols M–Q: M=12,N=13,O=14,P=15,Q=16) ──
    HT_MODELS_ROWS = range(2, 6)   # rows 3–6 only (4 HT models)
    for ri in HT_MODELS_ROWS:
        model = rows_vals[ri][12]
        if model is None or not str(model).strip():
            continue
        qp10 = _v(ri, 13)   # Plate 16mm*10mm
        qp5  = _v(ri, 14)   # Plate 5mm
        qa50 = _v(ri, 15)   # Angle 50*6
        qss  = _v(ri, 16)   # SS Sheet 3mm
        p10_cost = qp10 * R['plate_10mm']
        p5_cost  = qp5  * R['plate_5mm2']
        a50_cost = qa50 * R['ang_50']
        ss_cost  = qss  * R['ss_sheet']
        fab_cost = p10_cost + p5_cost + a50_cost + ss_cost
        selling  = round(fab_cost * MARKUP, 2)
        records.append({
            "type": "HT", "model": str(model).strip(),
            "plate_16_10_qty": qp10, "plate_16_10_cost": round(p10_cost, 2),
            "plate_5_qty":     qp5,  "plate_5_cost":     round(p5_cost, 2),
            "ang_50_qty":      qa50, "ang_50_cost":      round(a50_cost, 2),
            "ss_sheet_qty":    qss,  "ss_sheet_cost":    round(ss_cost, 2),
            "fabrication_cost": round(fab_cost, 2), "selling_price": selling,
        })

    # ── Type FS (rows 21–27, 0-indexed 20–26) ────────────────────
    # cols: A=0,B=1(tube_c),C=2(tube_b),D=3(ci_gills),E=4,F=5,G=6,H=7(bolt),I=8(channel)
    FS_MODELS_ROWS = range(20, 27)
    for ri in FS_MODELS_ROWS:
        model = rows_vals[ri][0]
        if model is None or not str(model).strip():
            continue
        qtc = _v(ri, 1)   # tube C qty (mtr)
        qtb = _v(ri, 2)   # tube B qty (mtr)
        qa6 = _v(ri, 4)   # Angle 65,50 (kg)
        qa1 = _v(ri, 5)   # Angle 100,100 (kg)
        qpl = _v(ri, 6)   # Plate 16mm*5mm (kg)
        qbl = _v(ri, 7)   # Hardware Bolt (kg)
        qch = _v(ri, 8)   # MS Channel (kg) — included in FS sum
        tube_c_cost = qtc * 4.27 * R['tube_c']
        tube_b_cost = qtb * 4.27 * R['tube_b']
        ang65_cost  = qa6 * R['ang_6550']
        ang100_cost = qa1 * R['ang_100']
        plate_cost  = qpl * R['plate_5mm']
        bolt_cost   = qbl * R['bolt']
        chan_cost   = qch * R['channel']
        fab_cost    = (tube_c_cost + tube_b_cost + ang65_cost + ang100_cost +
                       plate_cost + bolt_cost + chan_cost)
        selling     = round(fab_cost * MARKUP, 2)
        records.append({
            "type": "FS", "model": str(model).strip(),
            "ms_tube_c_qty": qtc, "ms_tube_c_cost": round(tube_c_cost, 2),
            "ms_tube_b_qty": qtb, "ms_tube_b_cost": round(tube_b_cost, 2),
            "ang_6550_qty":  qa6, "ang_6550_cost":  round(ang65_cost, 2),
            "ang_100100_qty":qa1, "ang_100100_cost": round(ang100_cost, 2),
            "plate_16_5_qty":qpl, "plate_16_5_cost": round(plate_cost, 2),
            "bolt_qty":      qbl, "bolt_cost":       round(bolt_cost, 2),
            "channel_qty":   qch, "channel_cost":    round(chan_cost, 2),
            "fabrication_cost": round(fab_cost, 2), "selling_price": selling,
        })

    pd.DataFrame(records).to_sql("recuperator_master", conn, if_exists="replace", index=False)
    return {"rows": len(records)}


# ─────────────────────────────────────────────────────────────────
# 9. GAIL GAS BURNER → gail_gas_burner_master
# ─────────────────────────────────────────────────────────────────

def parse_gail_gas_burner(xl, conn):
    sheet = _find_sheet(xl, "gail")
    if sheet is None:
        return {"skipped": "GAIL GAS Burner sheet not found"}

    df = xl.parse(sheet, header=None).dropna(how="all")
    records = []
    current_section = None
    headers = None

    for _, row in df.iterrows():
        non_null = [str(x).strip() for x in row if pd.notna(x) and str(x).strip() not in ("nan", "")]
        if not non_null:
            continue
        first = non_null[0]

        if "GAIL GAS BURNER" in first.upper() or "PILOT BURNER" in first.upper() or "ALL BURNER" in first.upper():
            current_section = first.upper().strip()
            headers = None
            continue

        if "BURNER SIZE" in first.upper():
            headers = non_null
            continue

        if headers and current_section and "ENCON" in first.upper():
            rec = {"section": current_section, "burner_size": first.upper()}
            for i, h in enumerate(headers[1:], 1):
                if i < len(non_null):
                    col = h.lower().replace(" ", "_").replace(".", "").replace("(", "").replace(")", "")
                    rec[col] = safe_float(non_null[i])
            records.append(rec)
        elif current_section and "PILOT" in current_section and len(non_null) >= 2:
            price = safe_float(non_null[1])
            if price is not None:
                records.append({"section": current_section, "burner_size": first.upper(), "burner": price})

    pd.DataFrame(records).to_sql("gail_gas_burner_master", conn, if_exists="replace", index=False)
    return {"rows": len(records)}


# ─────────────────────────────────────────────────────────────────
# 10. RAD HEAT → rad_heat_master / rad_heat_tata_master
# ─────────────────────────────────────────────────────────────────

def _parse_rad_heat(xl, sheet_name, table_name, conn):
    """
    Compute Rad Heat prices live from Rates C/G columns.

    Formula chain (verified for both 'Rad Heat' and 'Rad Heat (TATA)'):
      K{r} = H{r} * Rates!C{pipe_row} * Rates!G{pipe_row}  (SS tube cost)
      M{r} = I{r} * Rates!C{elbow_row}                     (SS elbow cost)
      O{r} = K{r} + M{r}
      E{r} = ROUND(E{base} + E{offset} + O{r}*4, -3)   (or without ROUND for TATA)

    Base/offset rows (E15–E20) are direct numbers, not formulas — unchanged on rate updates.
    """
    import openpyxl
    import re as _re

    sheet = _find_sheet(xl, sheet_name)
    if sheet is None:
        return {"skipped": f"'{sheet_name}' sheet not found"}

    # Live Rates C and G columns
    wb_rates = openpyxl.load_workbook(xl.io, read_only=True, data_only=True)
    _rs = next((s for s in wb_rates.sheetnames if s.strip().lower() == "rates"), None)
    if not _rs:
        wb_rates.close()
        return {"error": "Rates sheet not found"}
    ws_r = wb_rates[_rs]
    c_col = {}
    g_col = {}
    for _r in range(1, 45):
        cv = ws_r.cell(_r, 3).value
        gv = ws_r.cell(_r, 7).value
        try:
            if cv is not None: c_col[_r] = float(cv)
        except (ValueError, TypeError): pass
        try:
            if gv is not None: g_col[_r] = float(gv)
        except (ValueError, TypeError): pass
    wb_rates.close()

    wb_f = openpyxl.load_workbook(xl.io, read_only=True, data_only=False)
    wb_v = openpyxl.load_workbook(xl.io, read_only=True, data_only=True)
    # Exact match first (avoids 'rad heat' substring-matching 'Rad Heat (TATA)')
    _sn = next((s for s in wb_f.sheetnames
                if s.strip().lower() == sheet_name.lower()), None)
    if _sn is None:
        _sn = next((s for s in wb_f.sheetnames
                    if sheet_name.lower() in s.strip().lower()), None)
    if _sn is None:
        wb_f.close(); wb_v.close()
        return {"skipped": f"'{sheet_name}' sheet not found"}
    ws_f = wb_f[_sn]
    ws_v = wb_v[_sn]

    def _cell_val(row, col):
        """Direct value from formula workbook (int/float) or data_only workbook."""
        v = ws_f.cell(row, col).value
        if isinstance(v, (int, float)):
            return float(v)
        v2 = ws_v.cell(row, col).value
        try:
            return float(v2) if v2 is not None else None
        except (TypeError, ValueError):
            return None

    def _eval_cell(row, col, depth=0):
        """Recursively evaluate a formula cell to a float."""
        if depth > 10:
            return None
        v = ws_f.cell(row, col).value
        if v is None:
            return _cell_val(row, col)
        if isinstance(v, (int, float)):
            return float(v)
        if not isinstance(v, str) or not v.startswith('='):
            return _cell_val(row, col)

        expr = v[1:]  # strip '='

        # Extract ROUND wrapper
        m_rnd = _re.match(r'^ROUND\((.+),\s*(-?\d+)\)$', expr, _re.I)
        if m_rnd:
            expr, rnd_digits = m_rnd.group(1), int(m_rnd.group(2))
        else:
            rnd_digits = None

        # Substitute Rates!C{n} and Rates!G{n} first (before same-sheet subs)
        expr = _re.sub(r'Rates!\$?C\$?(\d+)',
                       lambda m: str(c_col.get(int(m.group(1)), 0.0)),
                       expr, flags=_re.I)
        expr = _re.sub(r'Rates!\$?G\$?(\d+)',
                       lambda m: str(g_col.get(int(m.group(1)), 0.0)),
                       expr, flags=_re.I)

        # Substitute remaining same-sheet cell refs (not preceded by '!')
        def _sub(m):
            if m.start() > 0 and expr[m.start() - 1] == '!':
                return m.group(0)
            rc = openpyxl.utils.column_index_from_string(m.group(1))
            rr = int(m.group(2))
            val = _eval_cell(rr, rc, depth + 1)
            return str(val) if val is not None else '0'

        expr = _re.sub(r'(?<!!)\b([A-Za-z]+)(\d+)\b', _sub, expr)

        if _re.match(r'^[\d\.\*\+\-/\(\)\s]+$', expr):
            try:
                result = eval(expr)
                if rnd_digits is not None:
                    return round(result, rnd_digits)
                return round(result, 2)
            except Exception:
                pass
        return None

    records = []
    section = None

    for r in range(1, 40):
        first_v = ws_v.cell(r, 1).value
        if first_v is None:
            continue
        first = str(first_v).strip()

        if "PRICE LIST OF RAD-HEAT" in first.upper():
            section = "models"
            continue
        if "SPARES PRICE" in first.upper():
            section = "spares"
            continue
        if first.upper().rstrip().rstrip('.') in ("MODEL", "ITEM", "MODEL "):
            continue

        if section == "models" and first.upper().startswith("ARE-"):
            output_kw = str(ws_v.cell(r, 2).value or '').strip() or None
            lpg = ws_v.cell(r, 3).value
            ng  = ws_v.cell(r, 4).value
            # E col (5) = price_with_ss_tubing — use cached Excel value
            cached = ws_v.cell(r, 5).value
            try: price = float(cached) if cached is not None else None
            except (TypeError, ValueError): price = None
            def _sf(v):
                try: return float(v) if v is not None else None
                except (TypeError, ValueError): return None
            records.append({
                "section": "MODEL",
                "item": first.upper(),
                "output_kw": output_kw,
                "gas_lpg_m3hr": _sf(lpg),
                "gas_ng_m3hr":  _sf(ng),
                "price_with_ss_tubing": price,
                # Tubing spec columns (G-O)
                "tube_dia_mm":    _sf(ws_v.cell(r, 7).value),
                "tube_length_m":  _sf(ws_v.cell(r, 8).value),
                "num_elbows":     _sf(ws_v.cell(r, 9).value),
                "ms_tube_cost":   _sf(ws_v.cell(r, 10).value),
                "ss_tube_cost":   _sf(ws_v.cell(r, 11).value),
                "ss_elbow_cost":  _sf(ws_v.cell(r, 13).value),
                "ms_total_cost":  _sf(ws_v.cell(r, 14).value),
                "ss_total_cost":  _sf(ws_v.cell(r, 15).value),
            })

        elif section == "spares":
            # Spares use direct values — not formula-dependent
            vals = []
            for c in range(1, 12):
                vv = ws_v.cell(r, c).value
                if vv is not None:
                    vals.append(str(vv).strip())
            if len(vals) >= 2:
                price = next((safe_float(v) for v in vals[1:] if safe_float(v) is not None), None)
                if price is not None:
                    records.append({
                        "section": "SPARE",
                        "item": vals[0].upper(),
                        "output_kw": None,
                        "gas_lpg_m3hr": None,
                        "gas_ng_m3hr":  None,
                        "price_with_ss_tubing": price,
                    })

    wb_f.close()
    wb_v.close()
    pd.DataFrame(records).to_sql(table_name, conn, if_exists="replace", index=False)
    return {"rows": len(records)}


# ─────────────────────────────────────────────────────────────────
# MAIN: parse everything
# ─────────────────────────────────────────────────────────────────

def parse_all(file_path: str, conn: sqlite3.Connection) -> dict:
    """
    Run all parsers against the given pricebook file.
    Returns a dict mapping table_name → {rows: N} or {skipped: reason} or {error: msg}.
    """
    xl = pd.ExcelFile(file_path)
    results = {}

    parsers = [
        ("component_price_master",    lambda: parse_rates(xl, conn)),
        ("hpu_master",                lambda: parse_hpu(xl, conn)),
        ("pumping_unit_price",        lambda: rebuild_pumping_unit_price(conn)),
        ("burner_pricelist_master",   lambda: parse_burner(xl, conn)),
        ("blower_pricelist_master",   lambda: parse_blower(xl, conn)),
        ("horizontal_master",         lambda: parse_horizontal(xl, conn)),
        ("vertical_master",           lambda: parse_vertical(xl, conn)),
        ("oil_burner_parts_master",   lambda: parse_oil_burner_parts(xl, conn)),
        ("hv_oil_burner_parts_master",lambda: parse_hv_oil_burner_parts(xl, conn)),
        ("gas_burner_parts_master",   lambda: parse_gas_burner_parts(xl, conn)),
        ("recuperator_master",        lambda: parse_recuperator(xl, conn)),
        ("gail_gas_burner_master",    lambda: parse_gail_gas_burner(xl, conn)),
        ("rad_heat_tata_master",      lambda: _parse_rad_heat(xl, "rad heat (tata)", "rad_heat_tata_master", conn)),
        ("rad_heat_master",           lambda: _parse_rad_heat(xl, "rad heat", "rad_heat_master", conn)),
    ]

    for table, fn in parsers:
        try:
            results[table] = fn()
        except Exception as exc:
            results[table] = {"error": str(exc)}

    conn.commit()
    return results
