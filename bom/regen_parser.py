# coding: utf-8
"""
bom/regen_parser.py
-------------------
Parses the Regen Standard Costing workbook:
  - Per-KW costing BOM items    (regen_costing_items)
  - Burner sizing: dims/wts/costs (regen_sizing)
  - Nozzle sizing (air/fume/NG DN) (regen_nozzle_sizing)
  - Material rate assumptions     (regen_material_rates)
  - Pipe size tables (4 gas types) (regen_pipe_sizes)
  - Price list summary             (regen_pricelist)
"""

import openpyxl
import pandas as pd

KW_SHEETS = {
    500:  'Costing REGENBurner 500 kw',
    1000: 'Costing REGENBurner 1000 Kw',
    1500: 'Costing REGENBurner 1500 Kw',
    2000: 'Costing REGENBurner 2000 Kw',
    2500: 'Costing REGENBurner 2500 Kw',
    3000: 'Costing REGENBurner 3000 Kw',
    4500: 'Costing REGENBurner 4500 Kw',
    6000: 'Costing REGENBurner 6000 Kw',
}


def _v(x):
    """Convert to float, handling formatted strings like '\xa0 1,285.51\xa0'."""
    if isinstance(x, (int, float)):
        return float(x)
    if isinstance(x, str):
        cleaned = x.replace('\xa0', '').replace(',', '').strip()
        try:
            return float(cleaned)
        except ValueError:
            return None
    return None


def _s(x):
    return str(x).strip().replace('\xa0', '') if x is not None else ''


def parse_regen_costing(xl_path, conn):
    """
    Parse the Regen Standard Costing workbook.
    Writes six DB tables:
        regen_costing_items, regen_sizing, regen_nozzle_sizing,
        regen_material_rates, regen_pipe_sizes, regen_pricelist
    """
    wb = openpyxl.load_workbook(xl_path, read_only=True, data_only=True)

    # ── 1. Per-KW costing sheets ─────────────────────────────────────────────
    costing_records = []

    for kw, shname in KW_SHEETS.items():
        if shname not in wb.sheetnames:
            continue
        ws = wb[shname]
        all_rows = list(ws.iter_rows(min_row=1, max_row=60, values_only=True))

        cur_section = ''
        row_order = 0

        for row in all_rows[2:]:          # skip header rows 1-2
            if all(v is None for v in (row[:10] if len(row) >= 10 else row)):
                continue

            sno        = _s(row[0]) if row[0] is not None else ''
            desc       = _s(row[1]) if len(row) > 1 and row[1] is not None else ''
            size_val   = row[2] if len(row) > 2 else None
            qty        = _v(row[3]) if len(row) > 3 else None
            cost_unit  = _v(row[4]) if len(row) > 4 else None
            total_cost = _v(row[5]) if len(row) > 5 else None
            sp_unit    = _v(row[6]) if len(row) > 6 else None
            total_sp   = _v(row[7]) if len(row) > 7 else None

            if not desc:
                # TOTAL row: both col F and col H are large non-zero numbers
                if total_cost is not None and total_sp is not None:
                    costing_records.append({
                        'kw': kw, 'row_order': row_order, 'row_type': 'total',
                        'sno': '', 'section': '', 'description': 'TOTAL',
                        'size': '', 'qty': None,
                        'cost_per_unit': None, 'total_cost': total_cost,
                        'sp_per_unit': None,   'total_sp': total_sp,
                    })
                    row_order += 1
                # FINAL row: only col H has a value
                elif total_sp is not None and total_cost is None and cost_unit is None:
                    costing_records.append({
                        'kw': kw, 'row_order': row_order, 'row_type': 'final',
                        'sno': '', 'section': '', 'description': 'FINAL SELLING PRICE',
                        'size': '', 'qty': None,
                        'cost_per_unit': None, 'total_cost': None,
                        'sp_per_unit': None,   'total_sp': total_sp,
                    })
                    row_order += 1
                continue

            if qty is None and cost_unit is None and total_cost is None:
                cur_section = desc
                row_type = 'header'
            else:
                row_type = 'item'

            costing_records.append({
                'kw': kw, 'row_order': row_order, 'row_type': row_type,
                'sno': sno,
                'section': cur_section if row_type == 'item' else '',
                'description': desc,
                'size': _s(size_val) if size_val is not None else '',
                'qty': qty,
                'cost_per_unit': cost_unit,
                'total_cost':    total_cost,
                'sp_per_unit':   sp_unit,
                'total_sp':      total_sp,
            })
            row_order += 1

    pd.DataFrame(costing_records).to_sql(
        'regen_costing_items', conn, if_exists='replace', index=False
    )

    # ── 2. Burner Sizing sheet ────────────────────────────────────────────────
    sizing_records   = []
    nozzle_records   = []
    mr_records       = []

    if 'Burner Sizing and costing' in wb.sheetnames:
        ws    = wb['Burner Sizing and costing']
        srows = list(ws.iter_rows(min_row=1, max_row=65, values_only=True))

        # ── Nozzle sizing table (rows 2-7 in Excel = index 1-6)
        # Cols 12-19: burner_name, power_kw, dn_air_in, air_speed,
        #             dn_fume_out, fume_speed, dn_ng_in, ng_speed
        for row in srows[1:8]:
            if len(row) < 20 or row[12] is None:
                continue
            burner_name = _s(row[12])
            if not burner_name or burner_name in ('REGEN', 'burner'):
                continue
            nozzle_records.append({
                'burner_name':  burner_name,
                'power_kw':     _v(row[13]),
                'dn_air_in':    _v(row[14]),
                'air_speed_ms': _v(row[15]),
                'dn_fume_out':  _v(row[16]),
                'fume_speed_ms':_v(row[17]),
                'dn_ng_in':     _v(row[18]),
                'ng_speed_ms':  _v(row[19]),
            })

        pd.DataFrame(nozzle_records).to_sql(
            'regen_nozzle_sizing', conn, if_exists='replace', index=False
        )

        # ── Dimension + detail weight table (Excel rows 20-27 = index 19-26)
        # All 31 cols captured; col 29 = wt_grand_total (formatted string)
        dim = {}
        for row in srows[19:27]:
            kw = _v(row[0])
            if kw is None:
                continue
            def rc(i): return _v(row[i]) if len(row) > i else None
            dim[int(kw)] = {
                'shell_thick':             rc(1),  'retainer_thick':      rc(2),
                'refractory_thick':        rc(3),
                'dim_L':                   rc(4),  'dim_H':               rc(5),
                'dim_W':                   rc(6),  'bottom_h':            rc(7),
                'vol_total':               rc(8),  'vol_effective':       rc(9),
                'vol_refractory':          rc(10), 'density_castable':    rc(11),
                'wt_refractory_insulation':rc(12),
                'loose_density_balls':     rc(13),
                'vol_available_balls':     rc(14), 'balls_filling_pct':   rc(15),
                'wt_ceramic_balls_burner': rc(16),
                'wt_shell':                rc(17), 'wt_ss_plate':         rc(18),
                'wt_regen_total':          rc(19),
                'bb_dia_inner':            rc(20), 'bb_dia_outer':        rc(21),
                'bb_depth':                rc(22), 'wt_burner_block':     rc(23),
                'burner_length':           rc(24), 'burner_dia':          rc(25),
                'wt_burner_shell':         rc(26), 'wt_burner_refrac_detail': rc(27),
                'wt_burner_total':         rc(28),
                'wt_grand_total':          rc(29), 'bloom_approx_wt':     rc(30),
            }

        # ── Weight summary table (Excel rows 35-42 = index 34-41)
        wt = {}
        for row in srows[34:42]:
            kw = _v(row[3])
            if kw is None:
                continue
            wt[int(kw)] = {
                'wt_burner_ms':          _v(row[4]),  'wt_burner_refrac':        _v(row[5]),
                'wt_regen_ms':           _v(row[6]),  'wt_regen_ss':             _v(row[7]),
                'wt_regen_refrac':       _v(row[8]),  'wt_ceramic_balls':        _v(row[9]),
                'wt_burner_block_summary':_v(row[10]),'wt_total':                _v(row[11]),
            }

        # ── Material rates (Excel rows 47-50 = index 46-49)
        mat_labels = ['MS', 'SS', 'Refractory', 'Ceramic Balls']
        for i, row in enumerate(srows[46:50]):
            label = _s(row[2]) or mat_labels[i]
            mr_records.append({
                'material':      label,
                'wastage':       _v(row[3]),
                'material_cost': _v(row[4]),
                'labor_cost':    _v(row[5]),
            })
        pd.DataFrame(mr_records).to_sql(
            'regen_material_rates', conn, if_exists='replace', index=False
        )

        # ── Cost breakdown (Excel rows 54-62 = index 53-61)
        cs = {}
        for row in srows[53:62]:
            kw = _v(row[3])
            if kw is None:
                continue
            cs[int(kw)] = {
                'cost_burner_ms':    _v(row[4]),  'cost_burner_refrac': _v(row[5]),
                'cost_regen_ms':     _v(row[6]),  'cost_regen_ss':      _v(row[7]),
                'cost_regen_refrac': _v(row[8]),  'cost_ceramic_balls': _v(row[9]),
                'cost_burner_block': _v(row[10]), 'cost_total':         _v(row[11]),
            }

        for kw in sorted(dim.keys()):
            rec = {'kw': kw}
            rec.update(dim.get(kw, {}))
            rec.update(wt.get(kw,  {}))
            rec.update(cs.get(kw,  {}))
            rec['selling_price'] = next(
                (r['total_sp'] for r in costing_records
                 if r['kw'] == kw and r['row_type'] == 'final'), None
            )
            sizing_records.append(rec)

        pd.DataFrame(sizing_records).to_sql(
            'regen_sizing', conn, if_exists='replace', index=False
        )

    # ── 3. Burner Pipe Size sheet ─────────────────────────────────────────────
    pipe_records = []
    if 'Burner Pipe Size' in wb.sheetnames:
        ws   = wb['Burner Pipe Size']
        prows = list(ws.iter_rows(min_row=1, max_row=55, values_only=True))

        # Each gas-type section: header row has a title in col A, then column-labels row,
        # then data rows until blank.  Detect sections by col-A string containing "GAS" or "Gas".
        gas_sections = [
            (6,  16, 'Natural Gas (NG) 8600 Kcal/Nm³'),
            (19, 29, 'Blast Furnace Gas 720 Kcal/Nm³'),
            (30, 40, 'Coke Oven Gas 4000 Kcal/Nm³'),
            (41, 51, 'Producer Gas 1250 Kcal/Nm³'),
        ]
        # Cols (0-based): 0=burner_size_kw, 1=gas_flow_nm3hr, 2=air_flow_nm3hr,
        #                  3=total_flue_nm3hr, 4=dn_air_mm, 5=dn_gas_mm, 6=dn_flue_mm,
        #                  7=area_air_m2, 8=area_gas_m2, 9=area_flue_m2,
        #                  10=vel_gas_ms, 11=vel_air_ms, 12=vel_flue_ms
        for (start0, end0, gas_name) in gas_sections:
            for row in prows[start0:end0]:
                kw = _v(row[0])
                if kw is None:
                    continue
                pipe_records.append({
                    'gas_type':         gas_name,
                    'burner_size_kw':   int(kw),
                    'gas_flow_nm3hr':   _v(row[1]),
                    'air_flow_nm3hr':   _v(row[2]),
                    'flue_flow_nm3hr':  _v(row[3]),
                    'dn_air_mm':        _v(row[4]),
                    'dn_gas_mm':        _v(row[5]),
                    'dn_flue_mm':       _v(row[6]),
                    'area_air_m2':      _v(row[7]),
                    'area_gas_m2':      _v(row[8]),
                    'area_flue_m2':     _v(row[9]),
                    'vel_gas_ms':       _v(row[10]),
                    'vel_air_ms':       _v(row[11]),
                    'vel_flue_ms':      _v(row[12]),
                })

    pd.DataFrame(pipe_records).to_sql(
        'regen_pipe_sizes', conn, if_exists='replace', index=False
    )

    # ── 4. Price List summary sheet ───────────────────────────────────────────
    if 'Price List' in wb.sheetnames:
        ws = wb['Price List']
        pl_records = []
        for row in ws.iter_rows(min_row=5, max_row=13, values_only=True):
            if not row or row[2] is None:
                continue
            model     = _s(row[2])
            kw        = _v(row[3])
            price_std = _v(row[4])
            price_wog = _v(row[5])
            per_kw    = _v(row[6])
            if model and kw:
                pl_records.append({
                    'model': model, 'kw': int(kw),
                    'price_std_complete': price_std,
                    'price_wo_gas_train': price_wog,
                    'price_per_kw':       per_kw,
                })
        pd.DataFrame(pl_records).to_sql(
            'regen_pricelist', conn, if_exists='replace', index=False
        )

    wb.close()
    return {
        'costing_items':  len(costing_records),
        'sizing_rows':    len(sizing_records),
        'nozzle_rows':    len(nozzle_records),
        'pipe_size_rows': len(pipe_records),
    }
